"""FDX V2 API contract.

The V2 router is additive: the original `/api` routes remain available to the
current dashboard while clients migrate to the versioned, enveloped contract.
"""

from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import tempfile
import uuid
import zipfile
from datetime import date, datetime, timedelta, timezone

import xlrd
from fastapi import (
    APIRouter,
    Cookie,
    Depends,
    File,
    Form,
    Header,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from PIL import Image
from pydantic import BaseModel, EmailStr, Field
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from starlette.background import BackgroundTask

from .auth import (
    access_token,
    check_login_rate_limit,
    check_public_rate_limit,
    create_refresh_session,
    current_user,
    find_user_by_email,
    hash_password,
    hash_token,
    new_opaque_token,
    require_org_admin,
    require_org_member,
    require_super_admin,
    rotate_refresh_session,
    verify_password,
)
from .config import settings
from .database import get_db
from .face_index import refresh_enrollment_matches, unique_face_stats
from .integrations import (
    dependency_health,
    dispatch_email,
    ml_embedding,
    queue_email,
    storage,
)
from .models import (
    AuditLog,
    Consent,
    Delivery,
    EmailOutbox,
    Event,
    FaceDetection,
    FaceEnrollment,
    FaceMatch,
    GalleryExport,
    IdempotencyRecord,
    Organization,
    OrganizationType,
    OutboxEvent,
    Participant,
    ParticipantEnrollmentToken,
    ParticipantImport,
    PasswordResetToken,
    Photo,
    ProcessingJob,
    RefreshSession,
    StorageReservation,
    StorageUsageLedger,
    UniqueFace,
    UploadBatch,
    User,
    UserInvitation,
    UserRole,
    WebhookEvent,
    utcnow,
)
from .serializers import event_json, organization_json
from .workflow import event_media_readiness

router = APIRouter(prefix="/api/v2")
GB = 1024**3
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
EVENT_TRANSITIONS = {
    "DRAFT": {"ENROLLMENT_OPEN", "DELETION_PENDING"},
    "ENROLLMENT_OPEN": {"READY_FOR_UPLOAD", "DELETION_PENDING"},
    "READY_FOR_UPLOAD": {"UPLOADING", "PROCESSING", "DELETION_PENDING"},
    "UPLOADING": {"PROCESSING", "DELETION_PENDING"},
    "PROCESSING": {"REVIEW", "READY_TO_DELIVER", "DELETION_PENDING"},
    "REVIEW": {"READY_TO_DELIVER", "PROCESSING", "DELETION_PENDING"},
    "READY_TO_DELIVER": {"DELIVERING", "PROCESSING", "DELETION_PENDING"},
    "DELIVERING": {"DELIVERED", "READY_TO_DELIVER", "DELETION_PENDING"},
    "DELIVERED": {"ARCHIVED", "DELETION_PENDING"},
    "ARCHIVED": {"EXPIRED", "DELETION_PENDING"},
    "EXPIRED": {"DELETION_PENDING"},
    "DELETION_PENDING": {"DELETED"},
}


def ok(data=None, request: Request | None = None, **meta):
    return {
        "data": data,
        "meta": {
            "request_id": getattr(request.state, "request_id", None) if request else None,
            **meta,
        },
    }


def add_audit(
    db: Session,
    user: User | None,
    action: str,
    details: str,
    organization_id: str | None = None,
    level: str = "info",
) -> None:
    db.add(
        AuditLog(
            organization_id=organization_id if organization_id is not None else user.organization_id if user else None,
            actor_user_id=user.id if user else None,
            actor=user.email if user else "system",
            action=action,
            details=details,
            level=level,
        )
    )


def add_outbox(
    db: Session,
    event_type: str,
    aggregate_type: str,
    aggregate_id: str,
    payload: dict,
    organization_id: str | None,
    correlation_id: str,
) -> OutboxEvent:
    event = OutboxEvent(
        aggregate_type=aggregate_type,
        aggregate_id=aggregate_id,
        organization_id=organization_id,
        event_type=event_type,
        event_version=1,
        correlation_id=correlation_id,
        payload=payload,
    )
    db.add(event)
    return event


def request_id(request: Request) -> str:
    return request.state.request_id


def user_v2(user: User) -> dict:
    return {
        "id": user.id,
        "name": user.name,
        "email": user.email,
        "role": user.role.value,
        "organization_id": user.organization_id,
        "status": user.status.upper(),
    }


def set_refresh_cookie(response: Response, token: str) -> None:
    response.set_cookie(
        "fdx_refresh",
        token,
        max_age=settings.refresh_token_days * 86400,
        httponly=True,
        secure=settings.environment == "production",
        samesite="strict",
        path="/api/v2/auth",
    )


def clear_refresh_cookie(response: Response) -> None:
    response.delete_cookie(
        "fdx_refresh",
        path="/api/v2/auth",
        secure=settings.environment == "production",
        samesite="strict",
    )


def tenant_event(db: Session, user: User, event_id: str, lock: bool = False) -> Event:
    statement = select(Event).where(Event.id == event_id, Event.organization_id == user.organization_id)
    if lock:
        statement = statement.with_for_update()
    event = db.scalar(statement)
    if not event or event.status.upper() in {"DELETION_PENDING", "DELETED"}:
        raise HTTPException(status_code=404, detail="Event was not found")
    return event


def require_event_media_ready(db: Session, event_id: str) -> Event:
    """Lock the event and prevent result delivery until every photo is ready."""
    event = db.scalar(select(Event).where(Event.id == event_id).with_for_update())
    if not event:
        raise HTTPException(status_code=404, detail="Event was not found")
    total, ready = event_media_readiness(db, event_id)
    if total == 0 or ready != total:
        raise HTTPException(
            status_code=409,
            detail=f"Event photos are still processing ({ready}/{total} ready)",
        )
    return event


def pagination(page: int, page_size: int) -> tuple[int, int]:
    if page < 1 or page_size < 1 or page_size > 100:
        raise HTTPException(
            status_code=422,
            detail="page must be >= 1 and page_size must be between 1 and 100",
        )
    return (page - 1) * page_size, page_size


def reserve_idempotency(db: Session, user: User, key: str | None, body: str) -> IdempotencyRecord | None:
    if not key:
        return None
    digest = hashlib.sha256(body.encode()).hexdigest()
    record = db.scalar(
        select(IdempotencyRecord).where(IdempotencyRecord.user_id == user.id, IdempotencyRecord.key == key)
    )
    if record:
        if record.request_hash != digest:
            raise HTTPException(
                status_code=409,
                detail="Idempotency key was already used for a different request",
            )
        return record
    record = IdempotencyRecord(
        user_id=user.id,
        key=key,
        request_hash=digest,
        expires_at=utcnow() + timedelta(days=1),
    )
    db.add(record)
    db.flush()
    return record


class LoginInput(BaseModel):
    email: EmailStr
    password: str


class PasswordInput(BaseModel):
    password: str = Field(min_length=10)


class ForgotPasswordInput(BaseModel):
    email: EmailStr


class ResetPasswordInput(BaseModel):
    token: str
    password: str = Field(min_length=10)


class OrganizationInput(BaseModel):
    name: str = Field(min_length=2, max_length=180)
    organization_type: OrganizationType
    primary_email: EmailStr
    contact_name: str = ""
    phone: str = ""
    storage_limit_bytes: int = Field(default=100 * GB, gt=0)
    default_retention_days: int = Field(default=90, ge=1, le=3650)
    account_expires_at: date | None = None


class OrganizationUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=2, max_length=180)
    primary_email: EmailStr | None = None
    contact_name: str | None = None
    phone: str | None = None
    storage_limit_bytes: int | None = Field(default=None, gt=0)
    default_retention_days: int | None = Field(default=None, ge=1, le=3650)
    account_expires_at: date | None = None


class InviteUserInput(BaseModel):
    name: str = Field(min_length=2, max_length=120)
    email: EmailStr


class EventInput(BaseModel):
    name: str = Field(min_length=2, max_length=180)
    description: str = ""
    location: str = ""
    starts_at: datetime
    ends_at: datetime | None = None
    retention_days: int | None = Field(default=None, ge=1, le=3650)
    enrollment_opens_at: datetime | None = None
    enrollment_closes_at: datetime | None = None
    gallery_expires_at: datetime | None = None


class EventUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=2, max_length=180)
    description: str | None = None
    location: str | None = None
    starts_at: datetime | None = None
    ends_at: datetime | None = None
    retention_days: int | None = Field(default=None, ge=1, le=3650)
    enrollment_opens_at: datetime | None = None
    enrollment_closes_at: datetime | None = None
    gallery_expires_at: datetime | None = None


class ParticipantInput(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    email: EmailStr


class UploadBatchInput(BaseModel):
    expected_files: int = Field(ge=1, le=100_000)
    reserved_bytes: int = Field(gt=0)


class UploadObjectInput(BaseModel):
    filename: str = Field(min_length=1, max_length=260)
    content_type: str
    size_bytes: int = Field(gt=0)
    sha256: str = Field(pattern=r"^[a-fA-F0-9]{64}$")


class PresignInput(BaseModel):
    files: list[UploadObjectInput] = Field(min_length=1, max_length=1000)


class MultipartPartInput(BaseModel):
    part_number: int = Field(ge=1, le=10_000)
    etag: str = Field(min_length=1, max_length=200)


class CompleteMultipartInput(BaseModel):
    upload_id: str = Field(min_length=1, max_length=500)
    parts: list[MultipartPartInput] = Field(min_length=1, max_length=10_000)


class BulkInvitationInput(BaseModel):
    enrollment_status: list[str] = Field(default_factory=lambda: ["invited", "opened"])
    search: str | None = Field(default=None, max_length=120)


class EnrollmentUploadInput(BaseModel):
    filename: str = Field(min_length=1, max_length=260)
    content_type: str
    size_bytes: int = Field(gt=0)
    sha256: str = Field(pattern=r"^[a-fA-F0-9]{64}$")


class MatchReviewInput(BaseModel):
    decision: str


@router.post("/auth/login")
def login(
    payload: LoginInput,
    response: Response,
    request: Request,
    db: Session = Depends(get_db),
):
    email = str(payload.email).lower()
    check_login_rate_limit(request, email)
    user = find_user_by_email(db, email)
    if not user or not verify_password(payload.password, user.password_hash):
        add_audit(db, None, "auth.login.failed", "Invalid credentials", level="warning")
        db.commit()
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if user.status != "active" or (user.organization and user.organization.status != "active"):
        raise HTTPException(status_code=403, detail="Account is not active")
    if user.password_hash and not user.password_hash.startswith("$argon2id$"):
        user.password_hash = hash_password(payload.password)
    raw_refresh, session = create_refresh_session(db, user, request)
    user.last_active_at = utcnow()
    add_audit(db, user, "auth.login.succeeded", f"Session {session.id} issued")
    tokens = access_token(user, session.id)
    db.commit()
    set_refresh_cookie(response, raw_refresh)
    return ok(
        {
            "access_token": tokens["access_token"],
            "expires_in": tokens["expires_in"],
            "user": user_v2(user),
            "redirect_to": (
                "/admin"
                if user.role == UserRole.SUPER_ADMIN
                else "/collaborator"
                if user.role == UserRole.COLLABORATOR
                else "/organization"
            ),
        },
        request,
    )


@router.post("/auth/refresh")
def refresh(
    response: Response,
    request: Request,
    fdx_refresh: str | None = Cookie(default=None),
    db: Session = Depends(get_db),
):
    if not fdx_refresh:
        raise HTTPException(status_code=401, detail="Refresh session is required")
    user, raw_refresh, session = rotate_refresh_session(db, fdx_refresh, request)
    tokens = access_token(user, session.id)
    db.commit()
    set_refresh_cookie(response, raw_refresh)
    return ok(
        {
            "access_token": tokens["access_token"],
            "expires_in": tokens["expires_in"],
            "user": user_v2(user),
        },
        request,
    )


@router.post("/auth/logout", status_code=204)
def logout(
    response: Response,
    fdx_refresh: str | None = Cookie(default=None),
    db: Session = Depends(get_db),
):
    if fdx_refresh:
        session = db.scalar(select(RefreshSession).where(RefreshSession.refresh_token_hash == hash_token(fdx_refresh)))
        if session and not session.revoked_at:
            session.revoked_at = utcnow()
            user = db.get(User, session.user_id)
            add_audit(db, user, "auth.logout", f"Session {session.id} revoked")
            db.commit()
    clear_refresh_cookie(response)


@router.get("/auth/me")
def me(request: Request, user: User = Depends(current_user)):
    return ok(user_v2(user), request)


@router.post("/auth/forgot-password", status_code=202)
def forgot_password(payload: ForgotPasswordInput, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "forgot-password", str(payload.email).lower(), 5, 300)
    user = find_user_by_email(db, str(payload.email).lower())
    if user and user.status == "active":
        db.query(PasswordResetToken).filter(
            PasswordResetToken.user_id == user.id,
            PasswordResetToken.consumed_at.is_(None),
        ).delete()
        raw_token, token_hash = new_opaque_token()
        db.add(
            PasswordResetToken(
                user_id=user.id,
                token_hash=token_hash,
                expires_at=utcnow() + timedelta(minutes=settings.password_reset_minutes),
            )
        )
        url = f"{settings.frontend_url}/reset-password/{raw_token}"
        item = queue_email(
            db,
            user.organization_id,
            user.email,
            "Reset your FDX password",
            f"<p><a href='{url}'>Reset password</a></p>",
        )
        dispatch_email(db, item)
        add_audit(db, user, "auth.password_reset.requested", "Password reset requested")
        db.commit()
    return ok(
        {"message": "If the account exists, a password reset email has been queued."},
        request,
    )


@router.post("/auth/reset-password")
def reset_password(payload: ResetPasswordInput, request: Request, db: Session = Depends(get_db)):
    token = db.scalar(
        select(PasswordResetToken).where(PasswordResetToken.token_hash == hash_token(payload.token)).with_for_update()
    )
    if not token or token.consumed_at or token.expires_at <= utcnow():
        raise HTTPException(status_code=404, detail="Password reset token is invalid or expired")
    user = db.get(User, token.user_id)
    user.password_hash = hash_password(payload.password)
    token.consumed_at = utcnow()
    db.query(RefreshSession).filter(RefreshSession.user_id == user.id, RefreshSession.revoked_at.is_(None)).update(
        {"revoked_at": utcnow()}
    )
    add_audit(
        db,
        user,
        "auth.password_reset.completed",
        "Password changed and sessions revoked",
    )
    db.commit()
    return ok({"message": "Password reset completed."}, request)


@router.post("/auth/invitations/{token}/accept")
def accept_invitation(
    token: str,
    payload: PasswordInput,
    response: Response,
    request: Request,
    db: Session = Depends(get_db),
):
    invitation = db.scalar(
        select(UserInvitation).where(UserInvitation.token_hash == hash_token(token)).with_for_update()
    )
    if not invitation or invitation.accepted_at or invitation.revoked_at or invitation.expires_at <= utcnow():
        raise HTTPException(status_code=404, detail="Invitation is invalid or expired")
    user = db.get(User, invitation.user_id)
    user.password_hash = hash_password(payload.password)
    user.status = "active"
    invitation.accepted_at = utcnow()
    raw_refresh, session = create_refresh_session(db, user, request)
    add_audit(db, user, "user.invitation.accepted", f"Invitation {invitation.id} accepted")
    tokens = access_token(user, session.id)
    db.commit()
    set_refresh_cookie(response, raw_refresh)
    return ok(
        {
            "access_token": tokens["access_token"],
            "expires_in": tokens["expires_in"],
            "user": user_v2(user),
        },
        request,
    )


@router.get("/admin/dashboard")
def admin_dashboard(
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    def count(model, *filters):
        return db.scalar(select(func.count(model.id)).where(*filters)) or 0

    data = {
        "organizations_total": count(Organization),
        "organizations_active": count(Organization, Organization.status == "active"),
        "organizations_suspended": count(Organization, Organization.status == "suspended"),
        "organization_users": count(User, User.organization_id.is_not(None)),
        "events_total": count(Event),
        "photos_total": count(Photo),
        "storage_used_bytes": db.scalar(select(func.coalesce(func.sum(Organization.storage_used_bytes), 0))) or 0,
        "jobs_queued": count(
            ProcessingJob,
            ProcessingJob.status.in_(["queued", "QUEUED", "RETRY_SCHEDULED"]),
        ),
        "jobs_running": count(ProcessingJob, ProcessingJob.status.in_(["processing", "RUNNING"])),
        "jobs_failed": count(
            ProcessingJob,
            ProcessingJob.status.in_(["failed", "FAILED", "DEAD_LETTERED"]),
        ),
        "emails_sent": count(EmailOutbox, EmailOutbox.status == "sent"),
        "emails_failed": count(EmailOutbox, EmailOutbox.status == "failed"),
        "expiring_events": count(
            Event,
            Event.expires_at <= date.today() + timedelta(days=7),
            Event.status.notin_(["expired", "DELETED"]),
        ),
    }
    return ok(data, request)


@router.get("/admin/system-health")
def system_health(request: Request, _: User = Depends(require_super_admin)):
    services = dependency_health()
    return ok(
        {
            "status": "healthy" if all(item["status"] == "healthy" for item in services) else "degraded",
            "services": services,
        },
        request,
    )


@router.get("/admin/organizations")
def organizations(
    request: Request,
    page: int = 1,
    page_size: int = 50,
    search: str | None = None,
    status: str | None = None,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    offset, limit = pagination(page, page_size)
    filters = []
    if search:
        filters.append(Organization.name.ilike(f"%{search.strip()}%"))
    if status:
        filters.append(Organization.status == status.lower())
    total = db.scalar(select(func.count(Organization.id)).where(*filters)) or 0
    rows = db.scalars(
        select(Organization).where(*filters).order_by(Organization.created_at.desc()).offset(offset).limit(limit)
    ).all()
    return ok(
        [organization_json(db, row) for row in rows],
        request,
        page=page,
        page_size=page_size,
        total=total,
    )


@router.post("/admin/organizations", status_code=201)
def create_organization(
    payload: OrganizationInput,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = Organization(
        name=payload.name.strip(),
        type=payload.organization_type,
        contact_name=payload.contact_name.strip(),
        contact_email=str(payload.primary_email).lower(),
        phone=payload.phone.strip(),
        storage_limit_bytes=payload.storage_limit_bytes,
        retention_days=payload.default_retention_days,
        expires_at=payload.account_expires_at,
        status="active",
    )
    db.add(item)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Organization conflicts with an existing record") from exc
    add_audit(db, user, "organization.created", item.name, item.id)
    db.commit()
    return ok(organization_json(db, item), request)


@router.get("/admin/organizations/{organization_id}")
def organization_detail(
    organization_id: str,
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    data = organization_json(db, item)
    data["recent_audit"] = [
        {
            "id": row.id,
            "action": row.action,
            "details": row.details,
            "created_at": row.created_at.isoformat(),
        }
        for row in db.scalars(
            select(AuditLog).where(AuditLog.organization_id == item.id).order_by(AuditLog.created_at.desc()).limit(20)
        ).all()
    ]
    return ok(data, request)


@router.patch("/admin/organizations/{organization_id}")
def update_organization(
    organization_id: str,
    payload: OrganizationUpdate,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    mapping = {
        "primary_email": "contact_email",
        "default_retention_days": "retention_days",
        "account_expires_at": "expires_at",
    }
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(
            item,
            mapping.get(key, key),
            str(value).lower() if key == "primary_email" else value,
        )
    add_audit(db, user, "organization.updated", ", ".join(payload.model_fields_set), item.id)
    db.commit()
    return ok(organization_json(db, item), request)


def set_org_status(organization_id: str, target: str, request: Request, user: User, db: Session):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    item.status = target
    if target != "active":
        session_ids = select(User.id).where(User.organization_id == item.id)
        db.query(RefreshSession).filter(
            RefreshSession.user_id.in_(session_ids), RefreshSession.revoked_at.is_(None)
        ).update({"revoked_at": utcnow()}, synchronize_session=False)
    add_audit(db, user, f"organization.{target}", item.name, item.id)
    db.commit()
    return ok(organization_json(db, item), request)


@router.post("/admin/organizations/{organization_id}/suspend")
def suspend_organization(
    organization_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return set_org_status(organization_id, "suspended", request, user, db)


@router.post("/admin/organizations/{organization_id}/activate")
def activate_organization(
    organization_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return set_org_status(organization_id, "active", request, user, db)


@router.post("/admin/organizations/{organization_id}/schedule-deletion", status_code=202)
def schedule_organization_deletion(
    organization_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    item.status = "deletion_pending"
    correlation = request_id(request)
    add_outbox(
        db,
        "fdx.v2.retention.cleanup.requested",
        "organization",
        item.id,
        {"resource_type": "organization", "resource_id": item.id},
        item.id,
        correlation,
    )
    add_audit(db, user, "organization.deletion_scheduled", item.name, item.id)
    db.commit()
    return ok({"status": "DELETION_PENDING"}, request)


@router.get("/admin/organizations/{organization_id}/storage")
def organization_storage(
    organization_id: str,
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    reserved = (
        db.scalar(
            select(func.coalesce(func.sum(StorageReservation.bytes), 0)).where(
                StorageReservation.organization_id == item.id,
                StorageReservation.status == "RESERVED",
                StorageReservation.expires_at > utcnow(),
            )
        )
        or 0
    )
    return ok(
        {
            "storage_limit_bytes": item.storage_limit_bytes,
            "storage_used_bytes": item.storage_used_bytes,
            "storage_reserved_bytes": reserved,
            "storage_available_bytes": max(0, item.storage_limit_bytes - item.storage_used_bytes - reserved),
        },
        request,
    )


class StoragePolicyInput(BaseModel):
    storage_limit_bytes: int = Field(gt=0)


@router.put("/admin/organizations/{organization_id}/storage")
def update_organization_storage(
    organization_id: str,
    payload: StoragePolicyInput,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    if payload.storage_limit_bytes < item.storage_used_bytes:
        raise HTTPException(status_code=422, detail="Storage limit cannot be below current usage")
    item.storage_limit_bytes = payload.storage_limit_bytes
    add_audit(
        db,
        user,
        "organization.storage_policy.updated",
        str(payload.storage_limit_bytes),
        item.id,
    )
    db.commit()
    return ok(
        {
            "storage_limit_bytes": item.storage_limit_bytes,
            "storage_used_bytes": item.storage_used_bytes,
        },
        request,
    )


@router.get("/admin/organizations/{organization_id}/retention")
def organization_retention(
    organization_id: str,
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    return ok(
        {
            "default_retention_days": item.retention_days,
            "account_expires_at": item.expires_at.isoformat() if item.expires_at else None,
        },
        request,
    )


class RetentionPolicyInput(BaseModel):
    default_retention_days: int = Field(ge=1, le=3650)
    account_expires_at: date | None = None


@router.put("/admin/organizations/{organization_id}/retention")
def update_organization_retention(
    organization_id: str,
    payload: RetentionPolicyInput,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = db.get(Organization, organization_id)
    if not item:
        raise HTTPException(status_code=404, detail="Organization was not found")
    item.retention_days = payload.default_retention_days
    item.expires_at = payload.account_expires_at
    add_audit(
        db,
        user,
        "organization.retention_policy.updated",
        f"{payload.default_retention_days} days",
        item.id,
    )
    db.commit()
    return ok(
        {
            "default_retention_days": item.retention_days,
            "account_expires_at": item.expires_at.isoformat() if item.expires_at else None,
        },
        request,
    )


@router.get("/admin/organizations/{organization_id}/users")
def organization_users(
    organization_id: str,
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not db.get(Organization, organization_id):
        raise HTTPException(status_code=404, detail="Organization was not found")
    rows = db.scalars(
        select(User).where(User.organization_id == organization_id).order_by(User.created_at.desc())
    ).all()
    return ok([user_v2(row) for row in rows], request)


@router.post("/admin/organizations/{organization_id}/users", status_code=201)
def invite_organization_user(
    organization_id: str,
    payload: InviteUserInput,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    organization = db.get(Organization, organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organization was not found")
    email = str(payload.email).lower()
    if find_user_by_email(db, email):
        raise HTTPException(status_code=409, detail="Email is already registered")
    invited = User(
        organization_id=organization.id,
        name=payload.name.strip(),
        email=email,
        role=UserRole.ORG_ADMIN,
        status="invited",
    )
    db.add(invited)
    db.flush()
    raw_token, token_hash = new_opaque_token()
    invitation = UserInvitation(
        user_id=invited.id,
        token_hash=token_hash,
        expires_at=utcnow() + timedelta(hours=settings.invitation_token_hours),
    )
    db.add(invitation)
    url = f"{settings.frontend_url}/accept-invite/{raw_token}"
    mail = queue_email(
        db,
        organization.id,
        invited.email,
        f"Join {organization.name} on FDX",
        f"<p><a href='{url}'>Set your password</a></p>",
    )
    dispatch_email(db, mail)
    add_audit(db, user, "user.invited", invited.email, organization.id)
    db.commit()
    data = user_v2(invited)
    if settings.environment == "development":
        data["development_invitation_url"] = url
    return ok(data, request)


def admin_user_or_404(db: Session, user_id: str) -> User:
    item = db.get(User, user_id)
    if not item or item.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=404, detail="Organization user was not found")
    return item


@router.get("/admin/users/{user_id}")
def admin_user_detail(
    user_id: str,
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return ok(user_v2(admin_user_or_404(db, user_id)), request)


class UserUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=2, max_length=120)
    email: EmailStr | None = None


@router.patch("/admin/users/{user_id}")
def admin_update_user(
    user_id: str,
    payload: UserUpdate,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = admin_user_or_404(db, user_id)
    if payload.name is not None:
        item.name = payload.name.strip()
    if payload.email is not None:
        item.email = str(payload.email).lower()
    add_audit(db, user, "user.updated", item.email, item.organization_id)
    db.commit()
    return ok(user_v2(item), request)


def set_user_status(user_id: str, target: str, request: Request, actor: User, db: Session):
    item = admin_user_or_404(db, user_id)
    item.status = target
    if target != "active":
        db.query(RefreshSession).filter(RefreshSession.user_id == item.id, RefreshSession.revoked_at.is_(None)).update(
            {"revoked_at": utcnow()}
        )
    add_audit(db, actor, f"user.{target}", item.email, item.organization_id)
    db.commit()
    return ok(user_v2(item), request)


@router.post("/admin/users/{user_id}/suspend")
def admin_suspend_user(
    user_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return set_user_status(user_id, "suspended", request, user, db)


@router.post("/admin/users/{user_id}/activate")
def admin_activate_user(
    user_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return set_user_status(user_id, "active", request, user, db)


@router.post("/admin/users/{user_id}/resend-invite")
def admin_resend_invite(
    user_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    item = admin_user_or_404(db, user_id)
    if item.password_hash:
        raise HTTPException(status_code=409, detail="User has already activated the account")
    db.query(UserInvitation).filter(
        UserInvitation.user_id == item.id,
        UserInvitation.accepted_at.is_(None),
        UserInvitation.revoked_at.is_(None),
    ).update({"revoked_at": utcnow()})
    raw_token, token_hash = new_opaque_token()
    invitation = UserInvitation(
        user_id=item.id,
        token_hash=token_hash,
        expires_at=utcnow() + timedelta(hours=settings.invitation_token_hours),
    )
    db.add(invitation)
    url = f"{settings.frontend_url}/accept-invite/{raw_token}"
    mail = queue_email(
        db,
        item.organization_id,
        item.email,
        "Your FDX invitation",
        f"<p><a href='{url}'>Set your password</a></p>",
    )
    dispatch_email(db, mail)
    add_audit(db, user, "user.invitation.resent", item.email, item.organization_id)
    db.commit()
    data = {"status": "QUEUED"}
    if settings.environment == "development":
        data["development_invitation_url"] = url
    return ok(data, request)


@router.get("/admin/jobs")
def admin_jobs(
    request: Request,
    page: int = 1,
    page_size: int = 50,
    status: str | None = None,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    offset, limit = pagination(page, page_size)
    filters = [ProcessingJob.status == status] if status else []
    total = db.scalar(select(func.count(ProcessingJob.id)).where(*filters)) or 0
    rows = db.scalars(
        select(ProcessingJob).where(*filters).order_by(ProcessingJob.created_at.desc()).offset(offset).limit(limit)
    ).all()
    data = [
        {
            "id": row.id,
            "organization_id": row.organization_id,
            "event_id": row.event_id,
            "job_type": row.job_type,
            "status": row.status,
            "attempt": row.attempt,
            "max_attempts": row.max_attempts,
            "progress_current": row.progress_current,
            "progress_total": row.progress_total,
            "error": row.error,
        }
        for row in rows
    ]
    return ok(data, request, page=page, page_size=page_size, total=total)


@router.get("/admin/jobs/{job_id}")
def admin_job_detail(
    job_id: str,
    request: Request,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    row = db.get(ProcessingJob, job_id)
    if not row:
        raise HTTPException(status_code=404, detail="Processing job was not found")
    return ok(
        {
            "id": row.id,
            "organization_id": row.organization_id,
            "event_id": row.event_id,
            "photo_id": row.photo_id,
            "job_type": row.job_type,
            "status": row.status,
            "attempt": row.attempt,
            "max_attempts": row.max_attempts,
            "progress_current": row.progress_current,
            "progress_total": row.progress_total,
            "correlation_id": row.correlation_id,
            "worker": row.worker,
            "error": row.error,
            "heartbeat_at": row.heartbeat_at.isoformat() if row.heartbeat_at else None,
            "created_at": row.created_at.isoformat(),
            "started_at": row.started_at.isoformat() if row.started_at else None,
            "finished_at": row.completed_at.isoformat() if row.completed_at else None,
        },
        request,
    )


@router.post("/admin/jobs/{job_id}/retry", status_code=202)
def admin_retry_job(
    job_id: str,
    request: Request,
    user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    event_id = db.scalar(select(ProcessingJob.event_id).where(ProcessingJob.id == job_id))
    if event_id:
        db.scalar(select(Event.id).where(Event.id == event_id).with_for_update())
    job = db.scalar(select(ProcessingJob).where(ProcessingJob.id == job_id).with_for_update())
    if not job or job.status not in {"failed", "FAILED", "DEAD_LETTERED"}:
        raise HTTPException(status_code=409, detail="Job is not retryable")
    job.status = "queued"
    job.error = None
    job.next_attempt_at = utcnow()
    add_outbox(
        db,
        "fdx.v2.ml.process.requested",
        "processing_job",
        job.id,
        {"job_id": job.id, "media_id": job.photo_id},
        job.organization_id,
        request_id(request),
    )
    add_audit(db, user, "processing.retry", job.id, job.organization_id)
    db.commit()
    return ok({"job_id": job.id, "status": "QUEUED"}, request)


@router.get("/admin/logs")
def admin_logs(
    request: Request,
    page: int = 1,
    page_size: int = 50,
    _: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    offset, limit = pagination(page, page_size)
    total = db.scalar(select(func.count(AuditLog.id))) or 0
    rows = db.scalars(select(AuditLog).order_by(AuditLog.created_at.desc()).offset(offset).limit(limit)).all()
    return ok(
        [
            {
                "id": row.id,
                "organization_id": row.organization_id,
                "actor": row.actor,
                "action": row.action,
                "details": row.details,
                "level": row.level,
                "created_at": row.created_at.isoformat(),
            }
            for row in rows
        ],
        request,
        page=page,
        page_size=page_size,
        total=total,
    )


@router.get("/organization")
def current_organization(
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    return ok(organization_json(db, user.organization), request)


@router.get("/organization/dashboard")
def organization_dashboard(
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    events = db.scalars(
        select(Event).where(Event.organization_id == user.organization_id).order_by(Event.created_at.desc())
    ).all()
    data = {
        "organization": organization_json(db, user.organization),
        "events": [event_json(db, event) for event in events],
        "participants": db.scalar(
            select(func.count(Participant.id)).where(Participant.organization_id == user.organization_id)
        )
        or 0,
        "photos": db.scalar(select(func.count(Photo.id)).where(Photo.organization_id == user.organization_id)) or 0,
        "failed_jobs": db.scalar(
            select(func.count(ProcessingJob.id)).where(
                ProcessingJob.organization_id == user.organization_id,
                ProcessingJob.status.in_(["failed", "FAILED", "DEAD_LETTERED"]),
            )
        )
        or 0,
    }
    return ok(data, request)


@router.get("/organization/usage")
def organization_usage(
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    reserved = (
        db.scalar(
            select(func.coalesce(func.sum(StorageReservation.bytes), 0)).where(
                StorageReservation.organization_id == user.organization_id,
                StorageReservation.status == "RESERVED",
                StorageReservation.expires_at > utcnow(),
            )
        )
        or 0
    )
    return ok(
        {
            "used_bytes": user.organization.storage_used_bytes,
            "reserved_bytes": reserved,
            "limit_bytes": user.organization.storage_limit_bytes,
            "available_bytes": max(
                0,
                user.organization.storage_limit_bytes - user.organization.storage_used_bytes - reserved,
            ),
        },
        request,
    )


@router.get("/organization/logs")
def organization_logs(
    request: Request,
    page: int = 1,
    page_size: int = 50,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    offset, limit = pagination(page, page_size)
    total = db.scalar(select(func.count(AuditLog.id)).where(AuditLog.organization_id == user.organization_id)) or 0
    rows = db.scalars(
        select(AuditLog)
        .where(AuditLog.organization_id == user.organization_id)
        .order_by(AuditLog.created_at.desc())
        .offset(offset)
        .limit(limit)
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "actor": row.actor,
                "action": row.action,
                "details": row.details,
                "level": row.level,
                "created_at": row.created_at.isoformat(),
            }
            for row in rows
        ],
        request,
        page=page,
        page_size=page_size,
        total=total,
    )


@router.get("/events")
def list_events(
    request: Request,
    page: int = 1,
    page_size: int = 50,
    status: str | None = None,
    search: str | None = None,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    offset, limit = pagination(page, page_size)
    filters = [Event.organization_id == user.organization_id]
    if status:
        filters.append(Event.status == status)
    if search:
        filters.append(Event.name.ilike(f"%{search.strip()}%"))
    total = db.scalar(select(func.count(Event.id)).where(*filters)) or 0
    rows = db.scalars(select(Event).where(*filters).order_by(Event.created_at.desc()).offset(offset).limit(limit)).all()
    return ok(
        [event_json(db, row) for row in rows],
        request,
        page=page,
        page_size=page_size,
        total=total,
    )


@router.post("/events", status_code=201)
def create_event(
    payload: EventInput,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    retention = min(
        payload.retention_days or user.organization.retention_days,
        user.organization.retention_days,
    )
    expires_at = (payload.starts_at + timedelta(days=retention)).date()
    item = Event(
        organization_id=user.organization_id,
        name=payload.name.strip(),
        description=payload.description,
        location=payload.location,
        event_date=payload.starts_at.date(),
        starts_at=payload.starts_at,
        ends_at=payload.ends_at,
        retention_days=retention,
        expires_at=expires_at,
        enrollment_opens_at=payload.enrollment_opens_at,
        enrollment_closes_at=payload.enrollment_closes_at,
        gallery_expires_at=payload.gallery_expires_at,
        created_by=user.id,
        status="DRAFT",
    )
    db.add(item)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="An event with the same name and date already exists",
        ) from exc
    add_audit(db, user, "event.created", item.name)
    db.commit()
    return ok(event_json(db, item), request)


@router.get("/events/{event_id}")
def get_event(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    return ok(event_json(db, tenant_event(db, user, event_id)), request)


@router.patch("/events/{event_id}")
def update_event(
    event_id: str,
    payload: EventUpdate,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    item = tenant_event(db, user, event_id, lock=True)
    values = payload.model_dump(exclude_unset=True)
    if "retention_days" in values:
        values["retention_days"] = min(values["retention_days"], user.organization.retention_days)
    for key, value in values.items():
        setattr(item, key, value)
    if payload.starts_at:
        item.event_date = payload.starts_at.date()
    item.expires_at = (
        item.starts_at or datetime.combine(item.event_date, datetime.min.time(), timezone.utc)
    ).date() + timedelta(days=item.retention_days)
    add_audit(db, user, "event.updated", f"{item.name}: {', '.join(values)}")
    db.commit()
    return ok(event_json(db, item), request)


def transition_event(event_id: str, target: str, request: Request, user: User, db: Session):
    item = tenant_event(db, user, event_id, lock=True)
    current = item.status.upper()
    if target not in EVENT_TRANSITIONS.get(current, set()):
        raise HTTPException(
            status_code=409,
            detail=f"Event cannot transition from {current} to {target}",
        )
    item.status = target
    add_audit(db, user, "event.state_changed", f"{current} -> {target}")
    db.commit()
    return ok({"id": item.id, "status": target}, request)


@router.post("/events/{event_id}/open-enrollment")
def open_enrollment(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    return transition_event(event_id, "ENROLLMENT_OPEN", request, user, db)


@router.post("/events/{event_id}/close-enrollment")
def close_enrollment(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    return transition_event(event_id, "READY_FOR_UPLOAD", request, user, db)


@router.post("/events/{event_id}/archive")
def archive_event(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    return transition_event(event_id, "ARCHIVED", request, user, db)


@router.delete("/events/{event_id}", status_code=202)
def delete_event(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    item = tenant_event(db, user, event_id, lock=True)
    item.status = "DELETION_PENDING"
    add_outbox(
        db,
        "fdx.v2.retention.cleanup.requested",
        "event",
        item.id,
        {"resource_type": "event", "resource_id": item.id},
        user.organization_id,
        request_id(request),
    )
    add_audit(db, user, "event.deletion_scheduled", item.name)
    db.commit()
    return ok({"id": item.id, "status": "DELETION_PENDING"}, request)


def parse_participants(content: bytes, filename: str) -> tuple[list[dict], list[dict], int]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    rows: list[dict] = []
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        rows = [dict(row) for row in csv.DictReader(io.StringIO(text))]
    elif suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        values = list(workbook.active.iter_rows(values_only=True))
        if values:
            headers = [str(value or "").strip() for value in values[0]]
            rows = [dict(zip(headers, values_row)) for values_row in values[1:]]
    elif suffix == "xls":
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, column)).strip() for column in range(sheet.ncols)] if sheet.nrows else []
        rows = [dict(zip(headers, sheet.row_values(index))) for index in range(1, sheet.nrows)]
    else:
        raise HTTPException(status_code=422, detail="Participant file must be CSV, XLS, XLSX, or XLSM")
    valid, errors, seen = [], [], set()
    for index, raw in enumerate(rows, start=2):
        normalized = {str(key).strip().lower(): value for key, value in raw.items()}
        name, email = (
            str(normalized.get("name") or "").strip(),
            str(normalized.get("email") or "").strip().lower(),
        )
        row_errors = []
        if not name:
            row_errors.append("name is required")
        if "@" not in email or email.startswith("@") or email.endswith("@"):
            row_errors.append("email is invalid")
        if email in seen:
            row_errors.append("duplicate email in file")
        if row_errors:
            errors.append({"row": index, "name": name, "email": email, "errors": row_errors})
        else:
            seen.add(email)
            valid.append({"name": name, "email": email})
    return valid, errors, len(rows)


@router.post("/events/{event_id}/participant-imports", status_code=201)
def create_participant_import(
    event_id: str,
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id)
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=422, detail="Import file is empty")
    valid, errors, total = parse_participants(content, file.filename or "participants.csv")
    existing = set(
        db.scalars(
            select(Participant.email).where(
                Participant.event_id == event.id,
                Participant.email.in_([row["email"] for row in valid]),
            )
        ).all()
    )
    duplicates = [row for row in valid if row["email"] in existing]
    accepted = [row for row in valid if row["email"] not in existing]
    item = ParticipantImport(
        organization_id=user.organization_id,
        event_id=event.id,
        source_filename=file.filename or "participants.csv",
        status="READY",
        total_rows=total,
        valid_rows=len(accepted),
        invalid_rows=len(errors),
        duplicate_rows=len(duplicates),
        validation_report={"errors": errors, "duplicates": duplicates},
        normalized_rows=accepted,
        created_by=user.id,
    )
    db.add(item)
    db.flush()
    add_audit(
        db,
        user,
        "participant_import.validated",
        f"{item.id}: {len(accepted)} valid, {len(errors)} invalid, {len(duplicates)} duplicate",
    )
    db.commit()
    return ok(
        {
            "id": item.id,
            "status": item.status,
            "total_rows": total,
            "valid_rows": len(accepted),
            "invalid_rows": len(errors),
            "duplicate_rows": len(duplicates),
            "errors": errors,
            "duplicates": duplicates,
        },
        request,
    )


@router.get("/events/{event_id}/participant-imports")
def list_participant_imports(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    rows = db.scalars(
        select(ParticipantImport)
        .where(
            ParticipantImport.event_id == event_id,
            ParticipantImport.organization_id == user.organization_id,
        )
        .order_by(ParticipantImport.created_at.desc())
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "filename": row.source_filename,
                "status": row.status,
                "total_rows": row.total_rows,
                "valid_rows": row.valid_rows,
                "invalid_rows": row.invalid_rows,
                "duplicate_rows": row.duplicate_rows,
                "created_at": row.created_at.isoformat(),
            }
            for row in rows
        ],
        request,
    )


@router.get("/events/{event_id}/participant-imports/{import_id}")
def get_participant_import(
    event_id: str,
    import_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    row = db.scalar(
        select(ParticipantImport).where(
            ParticipantImport.id == import_id,
            ParticipantImport.event_id == event_id,
            ParticipantImport.organization_id == user.organization_id,
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="Participant import was not found")
    return ok(
        {
            "id": row.id,
            "status": row.status,
            "total_rows": row.total_rows,
            "valid_rows": row.valid_rows,
            "invalid_rows": row.invalid_rows,
            "duplicate_rows": row.duplicate_rows,
            "validation_report": row.validation_report,
        },
        request,
    )


@router.post("/events/{event_id}/participant-imports/{import_id}/confirm", status_code=201)
def confirm_participant_import(
    event_id: str,
    import_id: str,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id)
    record = reserve_idempotency(db, user, idempotency_key, f"confirm-import:{import_id}")
    if record and record.response_body:
        return record.response_body
    item = db.scalar(
        select(ParticipantImport)
        .where(
            ParticipantImport.id == import_id,
            ParticipantImport.event_id == event.id,
            ParticipantImport.organization_id == user.organization_id,
        )
        .with_for_update()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Participant import was not found")
    if item.status == "CONFIRMED":
        raise HTTPException(status_code=409, detail="Participant import is already confirmed")
    created, invitations = [], []
    for row in item.normalized_rows or []:
        raw_token, token_hash = new_opaque_token()
        participant = Participant(
            organization_id=user.organization_id,
            event_id=event.id,
            name=row["name"],
            email=row["email"],
            enrollment_status="invited",
            delivery_status="pending",
            enrollment_token_hash=token_hash,
            enrollment_expires_at=utcnow() + timedelta(days=settings.enrollment_token_days),
        )
        db.add(participant)
        db.flush()
        db.add(
            ParticipantEnrollmentToken(
                participant_id=participant.id,
                token_hash=token_hash,
                expires_at=participant.enrollment_expires_at,
            )
        )
        url = f"{settings.frontend_url}/enroll/{raw_token}"
        mail = queue_email(
            db,
            user.organization_id,
            participant.email,
            f"Find your photos from {event.name}",
            f"<p><a href='{url}'>Find My Photos</a></p>",
        )
        dispatch_email(db, mail)
        created.append(participant.id)
        if settings.environment == "development":
            invitations.append({"participant_id": participant.id, "url": url})
    item.status = "CONFIRMED"
    item.confirmed_at = utcnow()
    add_audit(
        db,
        user,
        "participant_import.confirmed",
        f"{item.id}: {len(created)} participants",
    )
    result = ok(
        {
            "import_id": item.id,
            "participants_created": len(created),
            "development_invitations": invitations,
        },
        request,
    )
    if record:
        record.response_status = 201
        record.response_body = result
    db.commit()
    return result


@router.get("/events/{event_id}/participants")
def participants(
    event_id: str,
    request: Request,
    page: int = 1,
    page_size: int = 50,
    status: str | None = None,
    search: str | None = None,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    offset, limit = pagination(page, page_size)
    filters = [
        Participant.event_id == event_id,
        Participant.organization_id == user.organization_id,
    ]
    if status:
        filters.append(Participant.enrollment_status == status)
    if search:
        filters.append((Participant.name.ilike(f"%{search}%")) | (Participant.email.ilike(f"%{search}%")))
    total = db.scalar(select(func.count(Participant.id)).where(*filters)) or 0
    rows = db.scalars(
        select(Participant).where(*filters).order_by(Participant.created_at.desc()).offset(offset).limit(limit)
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "name": row.name,
                "email": row.email,
                "enrollment_status": row.enrollment_status,
                "delivery_status": row.delivery_status,
                "created_at": row.created_at.isoformat(),
            }
            for row in rows
        ],
        request,
        page=page,
        page_size=page_size,
        total=total,
    )


def tenant_participant(db: Session, user: User, event_id: str, participant_id: str) -> Participant:
    tenant_event(db, user, event_id)
    item = db.scalar(
        select(Participant).where(
            Participant.id == participant_id,
            Participant.event_id == event_id,
            Participant.organization_id == user.organization_id,
        )
    )
    if not item:
        raise HTTPException(status_code=404, detail="Participant was not found")
    return item


@router.get("/events/{event_id}/participants/{participant_id}")
def participant_detail(
    event_id: str,
    participant_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    item = tenant_participant(db, user, event_id, participant_id)
    matches = (
        db.scalar(
            select(func.count(FaceMatch.id)).where(
                FaceMatch.participant_id == item.id,
                FaceMatch.state == "high",
            )
        )
        or 0
    )
    return ok(
        {
            "id": item.id,
            "name": item.name,
            "email": item.email,
            "enrollment_status": item.enrollment_status,
            "delivery_status": item.delivery_status,
            "matches": matches,
        },
        request,
    )


@router.patch("/events/{event_id}/participants/{participant_id}")
def update_participant(
    event_id: str,
    participant_id: str,
    payload: ParticipantInput,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    item = tenant_participant(db, user, event_id, participant_id)
    item.name = payload.name.strip()
    item.email = str(payload.email).lower()
    add_audit(db, user, "participant.updated", item.email)
    db.commit()
    return ok({"id": item.id, "name": item.name, "email": item.email}, request)


@router.delete("/events/{event_id}/participants/{participant_id}", status_code=204)
def delete_participant(
    event_id: str,
    participant_id: str,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    item = tenant_participant(db, user, event_id, participant_id)
    if item.enrollment:
        storage.delete(item.enrollment.storage_key)
        user.organization.storage_used_bytes = max(0, user.organization.storage_used_bytes - item.enrollment.size_bytes)
        db.add(
            StorageUsageLedger(
                organization_id=user.organization_id,
                event_id=event_id,
                operation="DELETE",
                bytes=-item.enrollment.size_bytes,
            )
        )
    add_audit(db, user, "participant.deleted", item.email)
    db.delete(item)
    db.commit()


def send_participant_invite(db: Session, participant: Participant, user: User) -> str:
    db.query(ParticipantEnrollmentToken).filter(
        ParticipantEnrollmentToken.participant_id == participant.id,
        ParticipantEnrollmentToken.consumed_at.is_(None),
        ParticipantEnrollmentToken.revoked_at.is_(None),
    ).update({"revoked_at": utcnow()})
    raw_token, token_hash = new_opaque_token()
    expires = utcnow() + timedelta(days=settings.enrollment_token_days)
    participant.enrollment_token_hash = token_hash
    participant.enrollment_expires_at = expires
    participant.enrollment_status = "invited"
    db.add(ParticipantEnrollmentToken(participant_id=participant.id, token_hash=token_hash, expires_at=expires))
    url = f"{settings.frontend_url}/enroll/{raw_token}"
    mail = queue_email(
        db,
        participant.organization_id,
        participant.email,
        f"Find your photos from {participant.event.name}",
        f"<p><a href='{url}'>Find My Photos</a></p>",
    )
    dispatch_email(db, mail)
    add_audit(db, user, "participant.invitation.sent", participant.email)
    return url


@router.post("/events/{event_id}/participants/{participant_id}/send-invite")
@router.post("/events/{event_id}/participants/{participant_id}/resend-invite")
def participant_send_invite(
    event_id: str,
    participant_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    item = tenant_participant(db, user, event_id, participant_id)
    url = send_participant_invite(db, item, user)
    db.commit()
    data = {"status": "QUEUED"}
    if settings.environment == "development":
        data["development_enrollment_url"] = url
    return ok(data, request)


@router.post("/events/{event_id}/participants/send-invites", status_code=202)
def participant_send_invites(
    event_id: str,
    payload: BulkInvitationInput,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    body = json.dumps(payload.model_dump(), sort_keys=True)
    idem = reserve_idempotency(db, user, idempotency_key, f"bulk-invite:{event_id}:{body}")
    if idem and idem.response_body:
        return idem.response_body
    filters = [
        Participant.event_id == event_id,
        Participant.organization_id == user.organization_id,
        Participant.enrollment_status.in_(payload.enrollment_status),
    ]
    if payload.search:
        term = f"%{payload.search.strip()}%"
        filters.append((Participant.name.ilike(term)) | (Participant.email.ilike(term)))
    rows = db.scalars(select(Participant).where(*filters).order_by(Participant.created_at).limit(10_000)).all()
    development_urls = []
    for participant in rows:
        url = send_participant_invite(db, participant, user)
        if settings.environment == "development":
            development_urls.append({"participant_id": participant.id, "url": url})
    result = ok(
        {
            "status": "QUEUED",
            "invitations_queued": len(rows),
            "development_invitations": development_urls,
        },
        request,
    )
    if idem:
        idem.response_status = 202
        idem.response_body = result
    add_audit(
        db,
        user,
        "participant.invitation.bulk_queued",
        f"{event_id}: {len(rows)} invitations",
    )
    db.commit()
    return result


@router.post("/events/{event_id}/participants", status_code=201)
def create_participant(
    event_id: str,
    payload: ParticipantInput,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id)
    raw_token, token_hash = new_opaque_token()
    item = Participant(
        organization_id=user.organization_id,
        event_id=event.id,
        name=payload.name.strip(),
        email=str(payload.email).lower(),
        enrollment_status="invited",
        delivery_status="pending",
        enrollment_token_hash=token_hash,
        enrollment_expires_at=utcnow() + timedelta(days=settings.enrollment_token_days),
    )
    db.add(item)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Participant email already exists in this event") from exc
    url = f"{settings.frontend_url}/enroll/{raw_token}"
    db.add(
        ParticipantEnrollmentToken(
            participant_id=item.id,
            token_hash=token_hash,
            expires_at=item.enrollment_expires_at,
        )
    )
    mail = queue_email(
        db,
        user.organization_id,
        item.email,
        f"Find your photos from {event.name}",
        f"<p><a href='{url}'>Find My Photos</a></p>",
    )
    dispatch_email(db, mail)
    add_audit(db, user, "participant.created", item.email)
    db.commit()
    data = {
        "id": item.id,
        "name": item.name,
        "email": item.email,
        "enrollment_status": item.enrollment_status,
    }
    if settings.environment == "development":
        data["development_enrollment_url"] = url
    return ok(data, request)


@router.post("/events/{event_id}/upload-batches", status_code=201)
def create_upload_batch(
    event_id: str,
    payload: UploadBatchInput,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id, lock=True)
    active_reserved = (
        db.scalar(
            select(func.coalesce(func.sum(StorageReservation.bytes), 0)).where(
                StorageReservation.organization_id == user.organization_id,
                StorageReservation.status == "RESERVED",
                StorageReservation.expires_at > utcnow(),
            )
        )
        or 0
    )
    if payload.reserved_bytes > settings.max_upload_bytes:
        raise HTTPException(status_code=413, detail="Upload batch exceeds the configured maximum")
    if (
        user.organization.storage_used_bytes + active_reserved + payload.reserved_bytes
        > user.organization.storage_limit_bytes
    ):
        raise HTTPException(status_code=413, detail="Organization storage quota would be exceeded")
    batch = UploadBatch(
        organization_id=user.organization_id,
        event_id=event.id,
        expected_files=payload.expected_files,
        reserved_bytes=payload.reserved_bytes,
        created_by=user.id,
        status="CREATED",
    )
    db.add(batch)
    db.flush()
    reservation = StorageReservation(
        organization_id=user.organization_id,
        event_id=event.id,
        upload_batch_id=batch.id,
        bytes=payload.reserved_bytes,
        status="RESERVED",
        expires_at=utcnow() + timedelta(minutes=settings.upload_reservation_minutes),
    )
    db.add(reservation)
    db.add(
        StorageUsageLedger(
            organization_id=user.organization_id,
            event_id=event.id,
            operation="RESERVE",
            bytes=payload.reserved_bytes,
        )
    )
    add_audit(db, user, "upload_batch.created", f"{batch.id}: {payload.reserved_bytes} bytes")
    db.commit()
    return ok(
        {
            "id": batch.id,
            "status": batch.status,
            "reserved_bytes": batch.reserved_bytes,
            "reservation_expires_at": reservation.expires_at.isoformat(),
        },
        request,
    )


@router.get("/events/{event_id}/upload-batches")
def upload_batches(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    rows = db.scalars(
        select(UploadBatch)
        .where(
            UploadBatch.event_id == event_id,
            UploadBatch.organization_id == user.organization_id,
        )
        .order_by(UploadBatch.created_at.desc())
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "status": row.status,
                "expected_files": row.expected_files,
                "uploaded_files": row.uploaded_files,
                "reserved_bytes": row.reserved_bytes,
                "committed_bytes": row.committed_bytes,
                "created_at": row.created_at.isoformat(),
                "completed_at": row.completed_at.isoformat() if row.completed_at else None,
            }
            for row in rows
        ],
        request,
    )


@router.get("/events/{event_id}/upload-batches/{batch_id}")
def upload_batch_detail(
    event_id: str,
    batch_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    row = db.scalar(
        select(UploadBatch).where(
            UploadBatch.id == batch_id,
            UploadBatch.event_id == event_id,
            UploadBatch.organization_id == user.organization_id,
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="Upload batch was not found")
    return ok(
        {
            "id": row.id,
            "status": row.status,
            "expected_files": row.expected_files,
            "uploaded_files": row.uploaded_files,
            "reserved_bytes": row.reserved_bytes,
            "committed_bytes": row.committed_bytes,
            "manifest": row.manifest,
        },
        request,
    )


@router.post("/events/{event_id}/upload-batches/{batch_id}/presign")
def presign_uploads(
    event_id: str,
    batch_id: str,
    payload: PresignInput,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    batch = db.scalar(
        select(UploadBatch)
        .where(
            UploadBatch.id == batch_id,
            UploadBatch.event_id == event_id,
            UploadBatch.organization_id == user.organization_id,
        )
        .with_for_update()
    )
    if not batch or batch.status not in {"CREATED", "UPLOADING"}:
        raise HTTPException(status_code=409, detail="Upload batch cannot accept files")
    existing_manifest = list(batch.manifest or [])
    existing_bytes = sum(item["size_bytes"] for item in existing_manifest)
    total = sum(item.size_bytes for item in payload.files)
    if existing_bytes + total > batch.reserved_bytes:
        raise HTTPException(status_code=413, detail="Files exceed reserved upload bytes")
    if len(existing_manifest) + len(payload.files) > batch.expected_files:
        raise HTTPException(status_code=409, detail="Files exceed the upload batch manifest count")
    existing_hashes = {item["sha256"] for item in existing_manifest}
    manifest, urls = [], []
    for item in payload.files:
        if item.content_type not in ALLOWED_IMAGE_TYPES:
            raise HTTPException(status_code=422, detail=f"Unsupported media type for {item.filename}")
        if item.size_bytes > settings.max_media_file_bytes:
            raise HTTPException(status_code=413, detail=f"Media file is too large: {item.filename}")
        expected_extension = {
            "image/jpeg": {".jpg", ".jpeg"},
            "image/png": {".png"},
            "image/webp": {".webp"},
        }[item.content_type]
        suffix = "." + item.filename.lower().rsplit(".", 1)[-1] if "." in item.filename else ""
        if suffix not in expected_extension:
            raise HTTPException(
                status_code=422,
                detail=f"Filename extension does not match media type: {item.filename}",
            )
        if item.sha256.lower() in existing_hashes:
            raise HTTPException(
                status_code=409,
                detail=f"Duplicate file in upload batch: {item.filename}",
            )
        media_id = str(uuid.uuid4())
        extension = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp"}[item.content_type]
        key = f"organizations/{user.organization_id}/events/{event_id}/media/original/{media_id}{extension}"
        record = {
            "media_id": media_id,
            "filename": item.filename,
            "content_type": item.content_type,
            "size_bytes": item.size_bytes,
            "sha256": item.sha256.lower(),
            "storage_key": key,
        }
        multipart = (
            storage.create_multipart_upload(key, item.content_type, item.size_bytes, settings.multipart_part_bytes)
            if item.size_bytes >= settings.multipart_threshold_bytes
            else None
        )
        if multipart:
            record["multipart_upload_id"] = multipart["upload_id"]
            urls.append(
                {
                    **record,
                    "multipart": True,
                    "part_size": multipart["part_size"],
                    "parts": multipart["parts"],
                    "complete_url": f"/v2/events/{event_id}/upload-batches/{batch.id}/objects/{media_id}/complete-multipart",
                }
            )
        else:
            upload_url = storage.presign_put(key, item.content_type)
            if not upload_url:
                upload_url = f"/api/v2/events/{event_id}/upload-batches/{batch.id}/objects/{media_id}"
            urls.append(
                {
                    **record,
                    "multipart": False,
                    "upload_url": upload_url,
                    "method": "PUT",
                    "headers": {"Content-Type": item.content_type},
                }
            )
        manifest.append(record)
        existing_hashes.add(item.sha256.lower())
    batch.manifest = [*existing_manifest, *manifest]
    batch.status = "UPLOADING"
    db.commit()
    return ok({"batch_id": batch.id, "files": urls, "expires_in": 900}, request)


@router.put("/events/{event_id}/upload-batches/{batch_id}/objects/{media_id}", status_code=204)
async def local_upload_object(
    event_id: str,
    batch_id: str,
    media_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    if settings.storage_backend == "s3":
        raise HTTPException(status_code=404, detail="Direct local upload endpoint is disabled")
    tenant_event(db, user, event_id)
    batch = db.scalar(
        select(UploadBatch).where(
            UploadBatch.id == batch_id,
            UploadBatch.event_id == event_id,
            UploadBatch.organization_id == user.organization_id,
        )
    )
    record = next((row for row in batch.manifest or [] if row["media_id"] == media_id), None) if batch else None
    if not record:
        raise HTTPException(status_code=404, detail="Upload object was not found")
    content = await request.body()
    if len(content) != record["size_bytes"] or hashlib.sha256(content).hexdigest() != record["sha256"]:
        raise HTTPException(
            status_code=422,
            detail="Uploaded object size or checksum does not match manifest",
        )
    storage.put(record["storage_key"], content, record["content_type"])


@router.post("/events/{event_id}/upload-batches/{batch_id}/objects/{media_id}/complete-multipart")
def complete_multipart_upload(
    event_id: str,
    batch_id: str,
    media_id: str,
    payload: CompleteMultipartInput,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    batch = db.scalar(
        select(UploadBatch)
        .where(
            UploadBatch.id == batch_id,
            UploadBatch.event_id == event_id,
            UploadBatch.organization_id == user.organization_id,
        )
        .with_for_update()
    )
    record = next((row for row in batch.manifest or [] if row["media_id"] == media_id), None) if batch else None
    if not record or not record.get("multipart_upload_id"):
        raise HTTPException(status_code=404, detail="Multipart upload was not found")
    if record["multipart_upload_id"] != payload.upload_id:
        raise HTTPException(status_code=409, detail="Multipart upload ID does not match the manifest")
    storage.complete_multipart_upload(
        record["storage_key"],
        payload.upload_id,
        [part.model_dump() for part in payload.parts],
    )
    record["multipart_completed"] = True
    batch.manifest = list(batch.manifest)
    db.commit()
    return ok({"media_id": media_id, "status": "UPLOADED"}, request)


@router.post("/events/{event_id}/upload-batches/{batch_id}/complete", status_code=202)
def complete_upload_batch(
    event_id: str,
    batch_id: str,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id, lock=True)
    idem = reserve_idempotency(db, user, idempotency_key, f"complete-upload:{batch_id}")
    if idem and idem.response_body:
        return idem.response_body
    batch = db.scalar(
        select(UploadBatch)
        .where(
            UploadBatch.id == batch_id,
            UploadBatch.event_id == event.id,
            UploadBatch.organization_id == user.organization_id,
        )
        .with_for_update()
    )
    if not batch or batch.status not in {"UPLOADING", "VERIFYING"}:
        raise HTTPException(status_code=409, detail="Upload batch cannot be completed")
    batch.status = "VERIFYING"
    committed, jobs = 0, []
    for record in batch.manifest or []:
        try:
            stat = storage.stat(record["storage_key"])
        except FileNotFoundError as exc:
            raise HTTPException(status_code=409, detail=f"Upload is incomplete: {record['filename']}") from exc
        if stat["size"] != record["size_bytes"]:
            raise HTTPException(status_code=422, detail=f"Uploaded size mismatch: {record['filename']}")
        if db.scalar(select(Photo.id).where(Photo.event_id == event.id, Photo.sha256 == record["sha256"])):
            storage.delete(record["storage_key"])
            continue
        photo = Photo(
            id=record["media_id"],
            organization_id=user.organization_id,
            event_id=event.id,
            filename=record["filename"],
            storage_key=record["storage_key"],
            content_type=record["content_type"],
            size_bytes=record["size_bytes"],
            sha256=record["sha256"],
            processing_status="queued",
        )
        job = ProcessingJob(
            organization_id=user.organization_id,
            event_id=event.id,
            photo_id=photo.id,
            job_type="ML_PROCESS",
            status="queued",
            correlation_id=request_id(request),
            max_attempts=5,
        )
        db.add_all([photo, job])
        db.flush()
        add_outbox(
            db,
            "fdx.v2.ml.process.requested",
            "processing_job",
            job.id,
            {"job_id": job.id, "media_id": photo.id},
            user.organization_id,
            request_id(request),
        )
        db.add(
            StorageUsageLedger(
                organization_id=user.organization_id,
                event_id=event.id,
                photo_id=photo.id,
                operation="ADD",
                bytes=photo.size_bytes,
            )
        )
        committed += photo.size_bytes
        jobs.append(job.id)
    reservation = db.scalar(
        select(StorageReservation).where(StorageReservation.upload_batch_id == batch.id).with_for_update()
    )
    if reservation:
        reservation.status = "COMMITTED"
    batch.committed_bytes = committed
    batch.uploaded_files = len(jobs)
    batch.status = "COMPLETE"
    batch.completed_at = utcnow()
    user.organization.storage_used_bytes += committed
    event.status = "PROCESSING"
    add_audit(
        db,
        user,
        "upload_batch.completed",
        f"{batch.id}: {len(jobs)} media, {committed} bytes",
    )
    result = ok(
        {
            "batch_id": batch.id,
            "status": batch.status,
            "media_created": len(jobs),
            "jobs": jobs,
        },
        request,
    )
    if idem:
        idem.response_status = 202
        idem.response_body = result
    db.commit()
    return result


@router.post("/events/{event_id}/upload-batches/{batch_id}/cancel")
def cancel_upload_batch(
    event_id: str,
    batch_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    batch = db.scalar(
        select(UploadBatch)
        .where(
            UploadBatch.id == batch_id,
            UploadBatch.event_id == event_id,
            UploadBatch.organization_id == user.organization_id,
        )
        .with_for_update()
    )
    if not batch or batch.status == "COMPLETE":
        raise HTTPException(status_code=409, detail="Upload batch cannot be cancelled")
    for record in batch.manifest or []:
        if record.get("multipart_upload_id") and not record.get("multipart_completed"):
            storage.abort_multipart_upload(record["storage_key"], record["multipart_upload_id"])
        storage.delete(record["storage_key"])
    reservation = db.scalar(select(StorageReservation).where(StorageReservation.upload_batch_id == batch.id))
    if reservation:
        reservation.status = "RELEASED"
        db.add(
            StorageUsageLedger(
                organization_id=user.organization_id,
                event_id=event_id,
                operation="RELEASE",
                bytes=reservation.bytes,
            )
        )
    batch.status = "CANCELLED"
    add_audit(db, user, "upload_batch.cancelled", batch.id)
    db.commit()
    return ok({"batch_id": batch.id, "status": batch.status}, request)


@router.get("/events/{event_id}/media")
def event_media(
    event_id: str,
    request: Request,
    page: int = 1,
    page_size: int = 50,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    offset, limit = pagination(page, page_size)
    total = (
        db.scalar(
            select(func.count(Photo.id)).where(
                Photo.event_id == event_id,
                Photo.organization_id == user.organization_id,
            )
        )
        or 0
    )
    rows = db.scalars(
        select(Photo)
        .where(Photo.event_id == event_id, Photo.organization_id == user.organization_id)
        .order_by(Photo.uploaded_at.desc())
        .offset(offset)
        .limit(limit)
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "filename": row.filename,
                "mime_type": row.content_type,
                "size_bytes": row.size_bytes,
                "sha256": row.sha256,
                "status": row.processing_status,
                "uploaded_at": row.uploaded_at.isoformat(),
            }
            for row in rows
        ],
        request,
        page=page,
        page_size=page_size,
        total=total,
    )


def tenant_photo(db: Session, user: User, event_id: str, media_id: str) -> Photo:
    tenant_event(db, user, event_id)
    item = db.scalar(
        select(Photo).where(
            Photo.id == media_id,
            Photo.event_id == event_id,
            Photo.organization_id == user.organization_id,
        )
    )
    if not item:
        raise HTTPException(status_code=404, detail="Media was not found")
    return item


@router.get("/events/{event_id}/media/{media_id}")
def media_detail(
    event_id: str,
    media_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    item = tenant_photo(db, user, event_id, media_id)
    return ok(
        {
            "id": item.id,
            "filename": item.filename,
            "mime_type": item.content_type,
            "size_bytes": item.size_bytes,
            "sha256": item.sha256,
            "status": item.processing_status,
            "download_url": storage.presign_get(item.storage_key) or f"/api/media/{item.id}",
        },
        request,
    )


@router.delete("/events/{event_id}/media/{media_id}", status_code=204)
def delete_media(
    event_id: str,
    media_id: str,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    item = tenant_photo(db, user, event_id, media_id)
    released = item.size_bytes + item.thumbnail_size_bytes
    storage.delete(item.storage_key)
    if item.thumbnail_storage_key:
        storage.delete(item.thumbnail_storage_key)
    user.organization.storage_used_bytes = max(0, user.organization.storage_used_bytes - released)
    db.add(
        StorageUsageLedger(
            organization_id=user.organization_id,
            event_id=event_id,
            operation="DELETE",
            bytes=-released,
        )
    )
    add_audit(db, user, "media.deleted", item.filename)
    db.delete(item)
    db.commit()


@router.post("/events/{event_id}/media/{media_id}/reprocess", status_code=202)
def reprocess_media(
    event_id: str,
    media_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id, lock=True)
    item = tenant_photo(db, user, event_id, media_id)
    job = ProcessingJob(
        organization_id=user.organization_id,
        event_id=event_id,
        photo_id=item.id,
        job_type="ML_PROCESS",
        status="queued",
        correlation_id=request_id(request),
    )
    db.add(job)
    db.flush()
    add_outbox(
        db,
        "fdx.v2.ml.process.requested",
        "processing_job",
        job.id,
        {"job_id": job.id, "media_id": item.id},
        user.organization_id,
        request_id(request),
    )
    item.processing_status = "queued"
    add_audit(db, user, "media.reprocess_requested", item.filename)
    db.commit()
    return ok({"job_id": job.id, "status": "QUEUED"}, request)


@router.post("/events/{event_id}/start-processing", status_code=202)
def start_processing(
    event_id: str,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id, lock=True)
    idem = reserve_idempotency(db, user, idempotency_key, f"start-processing:{event_id}")
    if idem and idem.response_body:
        return idem.response_body
    queued = db.scalars(
        select(ProcessingJob).where(
            ProcessingJob.event_id == event.id,
            ProcessingJob.status.in_(["queued", "RETRY_SCHEDULED"]),
        )
    ).all()
    if not queued:
        raise HTTPException(status_code=409, detail="No media is queued for processing")
    event.status = "PROCESSING"
    for job in queued:
        add_outbox(
            db,
            "fdx.v2.ml.process.requested",
            "processing_job",
            job.id,
            {"job_id": job.id, "media_id": job.photo_id},
            user.organization_id,
            request_id(request),
        )
    add_audit(db, user, "processing.started", f"{event.name}: {len(queued)} jobs")
    result = ok(
        {"event_id": event.id, "status": event.status, "jobs_queued": len(queued)},
        request,
    )
    if idem:
        idem.response_status = 202
        idem.response_body = result
    db.commit()
    return result


@router.post("/events/{event_id}/cancel-processing", status_code=202)
def cancel_processing(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id, lock=True)
    jobs = db.scalars(
        select(ProcessingJob)
        .where(
            ProcessingJob.event_id == event.id,
            ProcessingJob.organization_id == user.organization_id,
            ProcessingJob.status.in_(["queued", "RETRY_SCHEDULED", "processing"]),
        )
        .with_for_update()
    ).all()
    cancelled = 0
    cancellation_requested = 0
    for job in jobs:
        if job.status == "processing":
            job.status = "CANCEL_REQUESTED"
            cancellation_requested += 1
        else:
            job.status = "CANCELLED"
            job.completed_at = utcnow()
            cancelled += 1
            if job.photo_id:
                photo = db.get(Photo, job.photo_id)
                if photo:
                    photo.processing_status = "uploaded"
    if event.status.upper() == "PROCESSING":
        event.status = "READY_FOR_UPLOAD"
    add_audit(
        db,
        user,
        "processing.cancelled",
        f"{event.id}: {cancelled} cancelled, {cancellation_requested} cancellation requested",
    )
    db.commit()
    return ok(
        {
            "event_id": event.id,
            "cancelled": cancelled,
            "cancellation_requested": cancellation_requested,
        },
        request,
    )


@router.get("/events/{event_id}/processing")
def processing_summary(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id)
    statuses = dict(
        db.execute(
            select(ProcessingJob.status, func.count(ProcessingJob.id))
            .where(ProcessingJob.event_id == event.id)
            .group_by(ProcessingJob.status)
        ).all()
    )
    photos_total = db.scalar(select(func.count(Photo.id)).where(Photo.event_id == event.id)) or 0
    photos_processed = (
        db.scalar(select(func.count(Photo.id)).where(Photo.event_id == event.id, Photo.processing_status == "ready"))
        or 0
    )
    unique_faces = db.scalar(select(func.count(UniqueFace.id)).where(UniqueFace.event_id == event.id)) or 0
    detections = db.scalar(select(func.count(FaceDetection.id)).where(FaceDetection.event_id == event.id)) or 0
    identity_stats = unique_face_stats(db, event_id=event.id)
    return ok(
        {
            "event_state": event.status,
            "photos_total": photos_total,
            "photos_processed": photos_processed,
            "photos_failed": statuses.get("failed", 0) + statuses.get("DEAD_LETTERED", 0),
            "faces_detected": unique_faces,
            "unique_faces": unique_faces,
            "face_detections": detections,
            "matches_auto": identity_stats["high"],
            "matches_review": identity_stats["review"],
            "matches_unknown": identity_stats["low"],
            "progress_percent": round(photos_processed * 100 / photos_total) if photos_total else 0,
        },
        request,
    )


@router.get("/events/{event_id}/processing/jobs")
def processing_jobs(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    rows = db.scalars(
        select(ProcessingJob)
        .where(
            ProcessingJob.event_id == event_id,
            ProcessingJob.organization_id == user.organization_id,
        )
        .order_by(ProcessingJob.created_at.desc())
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "photo_id": row.photo_id,
                "job_type": row.job_type,
                "status": row.status,
                "attempt": row.attempt,
                "max_attempts": row.max_attempts,
                "progress_current": row.progress_current,
                "progress_total": row.progress_total,
                "error_code": "PROCESSING_FAILED" if row.error else None,
                "error_message": row.error,
                "queued_at": row.created_at.isoformat(),
                "started_at": row.started_at.isoformat() if row.started_at else None,
                "finished_at": row.completed_at.isoformat() if row.completed_at else None,
            }
            for row in rows
        ],
        request,
    )


@router.get("/events/{event_id}/processing/jobs/{job_id}")
def processing_job_detail(
    event_id: str,
    job_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    row = db.scalar(
        select(ProcessingJob).where(
            ProcessingJob.id == job_id,
            ProcessingJob.event_id == event_id,
            ProcessingJob.organization_id == user.organization_id,
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="Processing job was not found")
    return ok(
        {
            "id": row.id,
            "photo_id": row.photo_id,
            "job_type": row.job_type,
            "status": row.status,
            "attempt": row.attempt,
            "max_attempts": row.max_attempts,
            "progress_current": row.progress_current,
            "progress_total": row.progress_total,
            "error_message": row.error,
            "heartbeat_at": row.heartbeat_at.isoformat() if row.heartbeat_at else None,
        },
        request,
    )


@router.post("/events/{event_id}/processing/jobs/{job_id}/retry", status_code=202)
def retry_processing_job(
    event_id: str,
    job_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id, lock=True)
    row = db.scalar(
        select(ProcessingJob)
        .where(
            ProcessingJob.id == job_id,
            ProcessingJob.event_id == event_id,
            ProcessingJob.organization_id == user.organization_id,
        )
        .with_for_update()
    )
    if not row or row.status not in {"failed", "FAILED", "DEAD_LETTERED"}:
        raise HTTPException(status_code=409, detail="Processing job is not retryable")
    row.status = "queued"
    row.error = None
    row.next_attempt_at = utcnow()
    add_outbox(
        db,
        "fdx.v2.ml.process.requested",
        "processing_job",
        row.id,
        {"job_id": row.id, "media_id": row.photo_id},
        user.organization_id,
        request_id(request),
    )
    add_audit(db, user, "processing.retry", row.id)
    db.commit()
    return ok({"job_id": row.id, "status": "QUEUED"}, request)


@router.get("/events/{event_id}/matches")
def event_matches(
    event_id: str,
    request: Request,
    decision: str | None = None,
    participant_id: str | None = None,
    media_id: str | None = None,
    review_required: bool | None = None,
    minimum_score: float | None = None,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    filters = [
        FaceMatch.event_id == event_id,
        FaceMatch.organization_id == user.organization_id,
    ]
    if decision:
        filters.append(FaceMatch.state == decision)
    if participant_id:
        filters.append(FaceMatch.participant_id == participant_id)
    if media_id:
        filters.append(FaceMatch.detection.has(FaceDetection.photo_id == media_id))
    if review_required is not None:
        filters.append(FaceMatch.state == "review" if review_required else FaceMatch.state != "review")
    if minimum_score is not None:
        filters.append(FaceMatch.confidence >= minimum_score)
    rows = db.scalars(select(FaceMatch).where(*filters).order_by(FaceMatch.created_at.desc()).limit(500)).all()
    return ok(
        [
            {
                "id": row.id,
                "participant_id": row.participant_id,
                "media_id": row.detection.photo_id,
                "similarity_score": row.confidence,
                "second_best_score": row.second_best_score,
                "margin": row.margin,
                "decision": row.state,
                "decision_source": row.decision_source,
                "model_name": row.model_name,
                "model_version": row.model_version,
                "threshold_profile_version": row.threshold_profile_version,
            }
            for row in rows
        ],
        request,
    )


@router.get("/events/{event_id}/matches/{match_id}")
def match_detail(
    event_id: str,
    match_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    row = db.scalar(
        select(FaceMatch).where(
            FaceMatch.id == match_id,
            FaceMatch.event_id == event_id,
            FaceMatch.organization_id == user.organization_id,
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="Match was not found")
    photo = row.detection.photo
    enrollment = row.participant.enrollment if row.participant else None
    return ok(
        {
            "id": row.id,
            "participant": {
                "id": row.participant.id,
                "name": row.participant.name,
                "email": row.participant.email,
            }
            if row.participant
            else None,
            "media": {
                "id": photo.id,
                "filename": photo.filename,
                "url": storage.presign_get(photo.storage_key) or f"/api/media/{photo.id}",
            },
            "detection": {
                "box": row.detection.box,
                "landmarks": row.detection.landmarks,
                "face_width": row.detection.face_width,
                "face_height": row.detection.face_height,
                "quality_class": row.detection.quality_class,
                "detector_confidence": row.detection.detector_confidence,
            },
            "enrollment_reference": {
                "url": storage.presign_get(enrollment.storage_key),
                "quality_score": enrollment.quality_score,
            }
            if enrollment
            else None,
            "similarity_score": row.confidence,
            "second_best_score": row.second_best_score,
            "margin": row.margin,
            "decision": row.state,
            "decision_source": row.decision_source,
            "model_name": row.model_name,
            "model_version": row.model_version,
            "threshold_profile_version": row.threshold_profile_version,
        },
        request,
    )


def review_match(event_id: str, match_id: str, target: str, request: Request, user: User, db: Session):
    tenant_event(db, user, event_id)
    raise HTTPException(
        status_code=410,
        detail=(
            "Manual match review is disabled; matches are approved automatically only above the confidence threshold"
        ),
    )


@router.post("/events/{event_id}/matches/{match_id}/confirm")
def confirm_match(
    event_id: str,
    match_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    return review_match(event_id, match_id, "approved", request, user, db)


@router.post("/events/{event_id}/matches/{match_id}/reject")
def reject_match(
    event_id: str,
    match_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    return review_match(event_id, match_id, "rejected", request, user, db)


@router.post("/events/{event_id}/galleries/build", status_code=202)
def build_galleries(
    event_id: str,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id, lock=True)
    require_event_media_ready(db, event.id)
    idem = reserve_idempotency(db, user, idempotency_key, f"build-galleries:{event_id}")
    if idem and idem.response_body:
        return idem.response_body
    participant_ids = db.scalars(
        select(FaceMatch.participant_id)
        .where(
            FaceMatch.event_id == event.id,
            FaceMatch.organization_id == user.organization_id,
            FaceMatch.participant_id.is_not(None),
            FaceMatch.state == "high",
        )
        .distinct()
    ).all()
    created = 0
    for participant_id in participant_ids:
        delivery = db.scalar(
            select(Delivery).where(Delivery.event_id == event.id, Delivery.participant_id == participant_id)
        )
        if not delivery:
            _, placeholder_hash = new_opaque_token()
            delivery = Delivery(
                organization_id=user.organization_id,
                event_id=event.id,
                participant_id=participant_id,
                gallery_token_hash=placeholder_hash,
                status="ready",
                expires_at=datetime.combine(event.expires_at, datetime.min.time(), timezone.utc),
            )
            db.add(delivery)
            created += 1
        add_outbox(
            db,
            "fdx.v2.gallery.build.requested",
            "participant",
            participant_id,
            {"participant_id": participant_id, "event_id": event.id},
            user.organization_id,
            request_id(request),
        )
    event.status = "READY_TO_DELIVER"
    add_audit(
        db,
        user,
        "gallery.build_requested",
        f"{event.name}: {len(participant_ids)} participants",
    )
    result = ok(
        {
            "event_id": event.id,
            "galleries_ready": len(participant_ids),
            "galleries_created": created,
        },
        request,
    )
    if idem:
        idem.response_status = 202
        idem.response_body = result
    db.commit()
    return result


@router.get("/events/{event_id}/galleries")
def list_galleries(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    rows = db.scalars(
        select(Delivery)
        .where(
            Delivery.event_id == event_id,
            Delivery.organization_id == user.organization_id,
        )
        .order_by(Delivery.created_at.desc())
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "participant_id": row.participant_id,
                "participant_name": row.participant.name,
                "status": row.status.upper(),
                "access_expires_at": row.expires_at.isoformat(),
                "created_at": row.created_at.isoformat(),
            }
            for row in rows
        ],
        request,
    )


@router.get("/events/{event_id}/galleries/{gallery_id}")
def gallery_detail(
    event_id: str,
    gallery_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    row = db.scalar(
        select(Delivery).where(
            Delivery.id == gallery_id,
            Delivery.event_id == event_id,
            Delivery.organization_id == user.organization_id,
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="Gallery was not found")
    count = (
        db.scalar(
            select(func.count(FaceMatch.id)).where(
                FaceMatch.event_id == event_id,
                FaceMatch.participant_id == row.participant_id,
                FaceMatch.state == "high",
            )
        )
        or 0
    )
    return ok(
        {
            "id": row.id,
            "participant_id": row.participant_id,
            "participant_name": row.participant.name,
            "status": row.status.upper(),
            "photos": count,
            "access_expires_at": row.expires_at.isoformat(),
        },
        request,
    )


def deliver_gallery(db: Session, delivery: Delivery, user: User) -> str:
    require_event_media_ready(db, delivery.event_id)
    raw_token, token_hash = new_opaque_token()
    delivery.gallery_token_hash = token_hash
    delivery.status = "ready"
    gallery_url = f"{settings.frontend_url}/gallery/{raw_token}"
    count = (
        db.scalar(
            select(func.count(FaceMatch.id)).where(
                FaceMatch.event_id == delivery.event_id,
                FaceMatch.participant_id == delivery.participant_id,
                FaceMatch.state == "high",
            )
        )
        or 0
    )
    mail = queue_email(
        db,
        delivery.organization_id,
        delivery.participant.email,
        f"Your photos from {delivery.event.name} are ready",
        f"<p>We found {count} photos containing you.</p><p><a href='{gallery_url}'>View My Photos</a></p>",
        delivery_id=delivery.id,
    )
    dispatch_email(db, mail)
    add_audit(db, user, "gallery.delivered", delivery.participant.email)
    return gallery_url


@router.post("/events/{event_id}/deliveries/send", status_code=202)
def send_deliveries(
    event_id: str,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    event = tenant_event(db, user, event_id, lock=True)
    idem = reserve_idempotency(db, user, idempotency_key, f"send-deliveries:{event_id}")
    if idem and idem.response_body:
        return idem.response_body
    rows = db.scalars(
        select(Delivery).where(
            Delivery.event_id == event.id,
            Delivery.organization_id == user.organization_id,
        )
    ).all()
    development_urls = []
    for row in rows:
        url = deliver_gallery(db, row, user)
        add_outbox(
            db,
            "fdx.v2.email.send.requested",
            "delivery",
            row.id,
            {"delivery_id": row.id},
            user.organization_id,
            request_id(request),
        )
        if settings.environment == "development":
            development_urls.append({"participant_id": row.participant_id, "url": url})
    event.status = "DELIVERING" if rows else event.status
    result = ok(
        {
            "event_id": event.id,
            "deliveries_queued": len(rows),
            "development_gallery_urls": development_urls,
        },
        request,
    )
    if idem:
        idem.response_status = 202
        idem.response_body = result
    db.commit()
    return result


@router.get("/events/{event_id}/deliveries")
def list_deliveries(
    event_id: str,
    request: Request,
    user: User = Depends(require_org_member),
    db: Session = Depends(get_db),
):
    tenant_event(db, user, event_id)
    rows = db.scalars(
        select(Delivery)
        .where(
            Delivery.event_id == event_id,
            Delivery.organization_id == user.organization_id,
        )
        .order_by(Delivery.created_at.desc())
    ).all()
    return ok(
        [
            {
                "id": row.id,
                "participant_id": row.participant_id,
                "participant_name": row.participant.name,
                "status": row.status.upper(),
                "sent_at": row.sent_at.isoformat() if row.sent_at else None,
                "expires_at": row.expires_at.isoformat(),
            }
            for row in rows
        ],
        request,
    )


@router.post("/events/{event_id}/participants/{participant_id}/resend-results")
def resend_results(
    event_id: str,
    participant_id: str,
    request: Request,
    user: User = Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    tenant_participant(db, user, event_id, participant_id)
    delivery = db.scalar(
        select(Delivery).where(
            Delivery.event_id == event_id,
            Delivery.participant_id == participant_id,
            Delivery.organization_id == user.organization_id,
        )
    )
    if not delivery:
        raise HTTPException(status_code=409, detail="Participant gallery has not been built")
    url = deliver_gallery(db, delivery, user)
    db.commit()
    data = {"status": "QUEUED"}
    if settings.environment == "development":
        data["development_gallery_url"] = url
    return ok(data, request)


def enrollment_access(
    db: Session,
    token: str,
    *,
    lock: bool = False,
) -> tuple[ParticipantEnrollmentToken | None, Participant, datetime]:
    """Resolve a reusable enrollment link without widening participant access."""

    token_hash = hash_token(token)
    token_query = select(ParticipantEnrollmentToken).where(ParticipantEnrollmentToken.token_hash == token_hash)
    if lock:
        token_query = token_query.with_for_update()
    token_row = db.scalar(token_query)
    participant = (
        db.get(Participant, token_row.participant_id)
        if token_row
        else db.scalar(select(Participant).where(Participant.enrollment_token_hash == token_hash))
    )
    expires_at = token_row.expires_at if token_row else participant.enrollment_expires_at if participant else None
    if (
        not participant
        or participant.event.status.upper() in {"DELETION_PENDING", "DELETED", "EXPIRED"}
        or not expires_at
        or expires_at <= utcnow()
        or (token_row and token_row.revoked_at)
    ):
        raise HTTPException(status_code=404, detail="Enrollment link is invalid or expired")
    return token_row, participant, expires_at


def effective_consent_policy_version(participant: Participant) -> str:
    organization_version = participant.event.organization.privacy_notice_version
    return f"{settings.consent_policy_version}:org-{organization_version}"


def enrollment_photos(token: str, participant: Participant, db: Session) -> list[dict]:
    rows = db.scalars(
        select(FaceMatch).where(
            FaceMatch.participant_id == participant.id,
            FaceMatch.event_id == participant.event_id,
            FaceMatch.state == "high",
        )
    ).all()
    photos = {row.detection.photo.id: row.detection.photo for row in rows}
    return [
        {
            "id": photo.id,
            "filename": photo.filename,
            "thumbnail_url": f"/api/v2/public/enrollment/{token}/photos/{photo.id}/thumbnail",
            "download_url": f"/api/v2/public/enrollment/{token}/photos/{photo.id}/download",
        }
        for photo in sorted(photos.values(), key=lambda item: item.uploaded_at, reverse=True)
    ]


def enrollment_photo(
    db: Session,
    participant: Participant,
    photo_id: str,
) -> Photo:
    photo = db.scalar(
        select(Photo)
        .join(FaceDetection, FaceDetection.photo_id == Photo.id)
        .join(FaceMatch, FaceMatch.detection_id == FaceDetection.id)
        .where(
            Photo.id == photo_id,
            Photo.event_id == participant.event_id,
            FaceMatch.participant_id == participant.id,
            FaceMatch.state == "high",
        )
        .limit(1)
    )
    if not photo:
        raise HTTPException(status_code=404, detail="Photo not found")
    return photo


@router.get("/public/enrollment/{token}")
def public_enrollment(token: str, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "enrollment-read", token, 30)
    token_row, participant, expires_at = enrollment_access(db, token, lock=True)
    if token_row and not token_row.opened_at:
        token_row.opened_at = utcnow()
        if participant.enrollment_status != "verified":
            participant.enrollment_status = "opened"
        db.commit()
    photos = enrollment_photos(token, participant, db) if participant.enrollment_status == "verified" else []
    organization = participant.event.organization
    return ok(
        {
            "organization_name": organization.name,
            "event_name": participant.event.name,
            "participant_name": participant.name,
            "status": participant.enrollment_status,
            "expires_at": expires_at.isoformat(),
            "purpose": "Find event photographs containing you",
            "retention_days": participant.event.retention_days,
            "consent_policy_version": effective_consent_policy_version(participant),
            "participant_privacy_notice": organization.participant_privacy_notice,
            "privacy_contact_email": organization.privacy_contact_email or organization.contact_email,
            "matching_results": {"photos": len(photos)},
            "photos": photos,
        },
        request,
    )


@router.post("/public/enrollment/{token}/consent", status_code=201)
def enrollment_consent(
    token: str,
    request: Request,
    accepted: bool = Form(...),
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "enrollment-consent", token, 10, 300)
    _, participant, _ = enrollment_access(db, token)
    if not accepted:
        raise HTTPException(status_code=422, detail="Consent is required")
    consent = Consent(
        organization_id=participant.organization_id,
        event_id=participant.event_id,
        participant_id=participant.id,
        consent_type="face_enrollment",
        policy_version=effective_consent_policy_version(participant),
        accepted=True,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
    )
    db.add(consent)
    db.commit()
    return ok(
        {
            "consent_id": consent.id,
            "policy_version": consent.policy_version,
            "accepted_at": consent.accepted_at.isoformat(),
        },
        request,
    )


@router.post("/public/enrollment/{token}/upload-url")
def enrollment_upload_url(
    token: str,
    payload: EnrollmentUploadInput,
    request: Request,
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "enrollment-upload-url", token, 10, 300)
    token_row, participant, _ = enrollment_access(db, token, lock=True)
    if not token_row:
        raise HTTPException(status_code=404, detail="Enrollment link is invalid or expired")
    if payload.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=422, detail="A JPEG, PNG, or WEBP selfie is required")
    if payload.size_bytes > settings.max_enrollment_bytes:
        raise HTTPException(status_code=413, detail="Enrollment image exceeds the configured maximum")
    expected_extensions = {
        "image/jpeg": {".jpg", ".jpeg"},
        "image/png": {".png"},
        "image/webp": {".webp"},
    }
    suffix = "." + payload.filename.lower().rsplit(".", 1)[-1] if "." in payload.filename else ""
    if suffix not in expected_extensions[payload.content_type]:
        raise HTTPException(status_code=422, detail="Filename extension does not match the image type")
    if token_row.pending_storage_key:
        storage.delete(token_row.pending_storage_key)
    extension = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp"}[payload.content_type]
    key = f"organizations/{participant.organization_id}/events/{participant.event_id}/enrollment-pending/{token_row.id}/{uuid.uuid4()}{extension}"
    token_row.pending_storage_key = key
    token_row.pending_content_type = payload.content_type
    token_row.pending_size_bytes = payload.size_bytes
    token_row.pending_sha256 = payload.sha256.lower()
    upload_url = storage.presign_put(key, payload.content_type) or f"/api/v2/public/enrollment/{token}/upload"
    db.commit()
    return ok(
        {
            "upload_url": upload_url,
            "method": "PUT",
            "headers": {"Content-Type": payload.content_type},
            "expires_in": 900,
        },
        request,
    )


@router.put("/public/enrollment/{token}/upload", status_code=204)
async def enrollment_local_upload(token: str, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "enrollment-upload", token, 10, 300)
    if storage.s3:
        raise HTTPException(status_code=404, detail="Direct upload endpoint is unavailable")
    token_row, _, _ = enrollment_access(db, token, lock=True)
    if not token_row or not token_row.pending_storage_key:
        raise HTTPException(status_code=404, detail="Enrollment upload is invalid or expired")
    content = await request.body()
    if len(content) != token_row.pending_size_bytes or hashlib.sha256(content).hexdigest() != token_row.pending_sha256:
        raise HTTPException(status_code=422, detail="Enrollment upload does not match its manifest")
    storage.put(token_row.pending_storage_key, content, token_row.pending_content_type)
    db.commit()
    return Response(status_code=204)


@router.post("/public/enrollment/{token}/complete")
def complete_enrollment(
    token: str,
    request: Request,
    selfie: UploadFile | None = File(default=None),
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "enrollment-complete", token, 10, 300)
    token_row, participant, expires_at = enrollment_access(db, token, lock=True)
    consent = db.scalar(
        select(Consent)
        .where(
            Consent.participant_id == participant.id,
            Consent.accepted.is_(True),
            Consent.policy_version == effective_consent_policy_version(participant),
        )
        .order_by(Consent.accepted_at.desc())
    )
    if not consent:
        raise HTTPException(status_code=422, detail="Consent must be recorded before enrollment")
    if selfie:
        content = selfie.file.read()
        content_type = selfie.content_type or "application/octet-stream"
        filename = selfie.filename or "selfie.jpg"
    elif token_row and token_row.pending_storage_key:
        try:
            info = storage.stat(token_row.pending_storage_key)
            content, stored_type = storage.read(token_row.pending_storage_key)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=409, detail="Enrollment upload has not completed") from exc
        if (
            info["size"] != token_row.pending_size_bytes
            or hashlib.sha256(content).hexdigest() != token_row.pending_sha256
        ):
            raise HTTPException(status_code=422, detail="Enrollment upload does not match its manifest")
        content_type = token_row.pending_content_type or stored_type
        filename = token_row.pending_storage_key.rsplit("/", 1)[-1]
    else:
        raise HTTPException(status_code=422, detail="An enrollment upload is required")
    if not content or content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=422, detail="A JPEG, PNG, or WEBP selfie is required")
    if len(content) > settings.max_enrollment_bytes:
        raise HTTPException(status_code=413, detail="Enrollment image exceeds the configured maximum")
    try:
        with Image.open(io.BytesIO(content)) as image:
            if image.width * image.height > settings.max_image_pixels:
                raise ValueError("Enrollment image exceeds the pixel limit")
            image.verify()
        result = ml_embedding(content, filename, content_type)
        face_width = int(result["box"].get("x_max", 0) - result["box"].get("x_min", 0))
        face_height = int(result["box"].get("y_max", 0) - result["box"].get("y_min", 0))
        if (
            min(face_width, face_height) < settings.minimum_face_size
            or result["box"]["probability"] < settings.minimum_detector_confidence
        ):
            raise ValueError("Enrollment face is too small or unclear")
    except Exception as exc:
        raise HTTPException(status_code=422, detail="A clear, usable face could not be enrolled") from exc
    previous_size = participant.enrollment.size_bytes if participant.enrollment else 0
    additional = len(content) - previous_size
    if (
        participant.event.organization.storage_used_bytes + additional
        > participant.event.organization.storage_limit_bytes
    ):
        raise HTTPException(status_code=413, detail="Organization storage quota would be exceeded")
    extension = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp"}[content_type]
    key = f"organizations/{participant.organization_id}/events/{participant.event_id}/enrollment/{participant.id}/{uuid.uuid4()}{extension}"
    storage.put(key, content, content_type)
    previous_key = participant.enrollment.storage_key if participant.enrollment else None
    enrollment = participant.enrollment or FaceEnrollment(
        participant_id=participant.id,
        storage_key=key,
        embedding=result["embedding"],
        detector_confidence=result["box"]["probability"],
    )
    enrollment.organization_id = participant.organization_id
    enrollment.event_id = participant.event_id
    enrollment.storage_key = key
    enrollment.size_bytes = len(content)
    enrollment.embedding = result["embedding"]
    enrollment.embedding_vector = result["embedding"]
    enrollment.embedding_dimension = len(result["embedding"])
    enrollment.detector_confidence = result["box"]["probability"]
    enrollment.quality_score = min(1.0, min(face_width, face_height) / settings.low_resolution_face_size)
    enrollment.model_name = "adaface-ir101-ms1mv2"
    enrollment.model_version = settings.embedder_model_version
    enrollment.status = "valid"
    enrollment.expires_at = datetime.combine(participant.event.expires_at, datetime.min.time(), timezone.utc)
    db.add(enrollment)
    participant.enrollment_status = "verified"
    participant.consented_at = consent.accepted_at
    participant.event.organization.storage_used_bytes += additional
    db.add(
        StorageUsageLedger(
            organization_id=participant.organization_id,
            event_id=participant.event_id,
            operation="ADD",
            bytes=additional,
        )
    )
    if token_row:
        token_row.consumed_at = token_row.consumed_at or utcnow()
        pending_key = token_row.pending_storage_key
        token_row.pending_storage_key = None
        token_row.pending_content_type = None
        token_row.pending_size_bytes = None
        token_row.pending_sha256 = None
    matching_results = refresh_enrollment_matches(db, participant, result["embedding"])
    photos = enrollment_photos(token, participant, db)
    add_audit(db, None, "enrollment.completed", participant.email, participant.organization_id)
    db.commit()
    if previous_key and previous_key != key:
        storage.delete(previous_key)
    if token_row and pending_key and pending_key != key:
        storage.delete(pending_key)
    return ok(
        {
            "status": "ENROLLED",
            "embedding_dimension": enrollment.embedding_dimension,
            "model_name": enrollment.model_name,
            "model_version": enrollment.model_version,
            "matching_results": matching_results,
            "event_name": participant.event.name,
            "organization_name": participant.event.organization.name,
            "expires_at": expires_at.isoformat(),
            "photos": photos,
        },
        request,
    )


@router.get("/public/enrollment/{token}/photos/{photo_id}/thumbnail")
def enrollment_photo_thumbnail(
    token: str,
    photo_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "enrollment-photo", token, 120)
    _, participant, _ = enrollment_access(db, token)
    photo = enrollment_photo(db, participant, photo_id)
    key = photo.thumbnail_storage_key or photo.storage_key
    content, content_type = storage.read(key)
    return Response(
        content,
        media_type=content_type,
        headers={"Cache-Control": "private, max-age=300"},
    )


@router.get("/public/enrollment/{token}/photos/{photo_id}/download")
def enrollment_photo_download(
    token: str,
    photo_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "enrollment-photo-download", token, 60)
    _, participant, _ = enrollment_access(db, token)
    photo = enrollment_photo(db, participant, photo_id)
    content, content_type = storage.read(photo.storage_key)
    safe_filename = photo.filename.replace('"', "")
    return Response(
        content,
        media_type=content_type,
        headers={
            "Cache-Control": "private, no-store",
            "Content-Disposition": f'attachment; filename="{safe_filename}"',
        },
    )


class EnrollmentDownloadInput(BaseModel):
    media_ids: list[str] = Field(min_length=1, max_length=10_000)


@router.post("/public/enrollment/{token}/download")
def enrollment_selected_download(
    token: str,
    payload: EnrollmentDownloadInput,
    request: Request,
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "enrollment-selected-download", token, 5, 300)
    _, participant, _ = enrollment_access(db, token)
    requested_ids = list(dict.fromkeys(payload.media_ids))
    rows = db.scalars(
        select(Photo)
        .join(FaceDetection, FaceDetection.photo_id == Photo.id)
        .join(FaceMatch, FaceMatch.detection_id == FaceDetection.id)
        .where(
            Photo.id.in_(requested_ids),
            Photo.event_id == participant.event_id,
            FaceMatch.participant_id == participant.id,
            FaceMatch.state == "high",
        )
        .distinct()
    ).all()
    permitted = {photo.id: photo for photo in rows}
    if set(permitted) != set(requested_ids):
        raise HTTPException(status_code=404, detail="One or more selected photos were not found")

    archive_buffer = tempfile.SpooledTemporaryFile(max_size=32 * 1024 * 1024)
    with zipfile.ZipFile(archive_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, photo_id in enumerate(requested_ids, start=1):
            photo = permitted[photo_id]
            content, _ = storage.read(photo.storage_key)
            filename = photo.filename.replace("..", "_").replace("/", "_").replace("\\", "_")
            archive.writestr(f"{index:04d}-{photo.id[:8]}-{filename}", content)
    archive_buffer.seek(0)
    event_name = (
        "".join(
            character if character.isalnum() or character in "-_" else "-" for character in participant.event.name
        ).strip("-")
        or "fdx-event"
    )

    def chunks():
        while content := archive_buffer.read(1024 * 1024):
            yield content

    return StreamingResponse(
        chunks(),
        media_type="application/zip",
        headers={
            "Cache-Control": "private, no-store",
            "Content-Disposition": f'attachment; filename="{event_name}-photos.zip"',
        },
        background=BackgroundTask(archive_buffer.close),
    )


@router.get("/public/gallery/{token}")
def public_gallery(token: str, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "gallery-read", token, 60)
    delivery = db.scalar(select(Delivery).where(Delivery.gallery_token_hash == hash_token(token)))
    if (
        not delivery
        or delivery.event.status.upper() in {"DELETION_PENDING", "DELETED", "EXPIRED"}
        or delivery.expires_at <= utcnow()
    ):
        raise HTTPException(status_code=404, detail="Gallery link is invalid or expired")
    rows = db.scalars(
        select(FaceMatch).where(
            FaceMatch.participant_id == delivery.participant_id,
            FaceMatch.event_id == delivery.event_id,
            FaceMatch.state == "high",
        )
    ).all()
    photos = {row.detection.photo.id: row.detection.photo for row in rows}
    data = []
    for photo in photos.values():
        signed = storage.presign_get(photo.thumbnail_storage_key or photo.storage_key)
        data.append(
            {
                "id": photo.id,
                "filename": photo.filename,
                "thumbnail_url": signed or f"/api/public/gallery/{token}/photos/{photo.id}/thumbnail",
                "download_endpoint": f"/api/v2/public/gallery/{token}/download-url",
            }
        )
    return ok(
        {
            "event_name": delivery.event.name,
            "organization_name": delivery.event.organization.name,
            "expires_at": delivery.expires_at.isoformat(),
            "photos": data,
        },
        request,
    )


class DownloadInput(BaseModel):
    media_id: str


class GalleryExportInput(BaseModel):
    media_ids: list[str] | None = Field(default=None, max_length=10_000)


@router.post("/public/gallery/{token}/download-url")
def gallery_download_url(token: str, payload: DownloadInput, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "gallery-download", token, 60)
    delivery = db.scalar(select(Delivery).where(Delivery.gallery_token_hash == hash_token(token)))
    if (
        not delivery
        or delivery.event.status.upper() in {"DELETION_PENDING", "DELETED", "EXPIRED"}
        or delivery.expires_at <= utcnow()
    ):
        raise HTTPException(status_code=404, detail="Gallery link is invalid or expired")
    match = db.scalar(
        select(FaceMatch)
        .join(FaceDetection)
        .where(
            FaceMatch.participant_id == delivery.participant_id,
            FaceMatch.event_id == delivery.event_id,
            FaceDetection.photo_id == payload.media_id,
            FaceMatch.state == "high",
        )
    )
    if not match:
        raise HTTPException(status_code=404, detail="Photo was not found")
    photo = match.detection.photo
    url = (
        storage.presign_get(photo.storage_key, filename=photo.filename)
        or f"/api/public/gallery/{token}/photos/{photo.id}"
    )
    return ok({"url": url, "expires_in": 600}, request)


@router.post("/public/gallery/{token}/download-all", status_code=202)
@router.post("/public/gallery/{token}/exports", status_code=202)
def create_gallery_export(
    token: str,
    request: Request,
    payload: GalleryExportInput | None = None,
    db: Session = Depends(get_db),
):
    check_public_rate_limit(request, "gallery-export", token, 5, 300)
    delivery = db.scalar(select(Delivery).where(Delivery.gallery_token_hash == hash_token(token)))
    if (
        not delivery
        or delivery.event.status.upper() in {"DELETION_PENDING", "DELETED", "EXPIRED"}
        or delivery.expires_at <= utcnow()
    ):
        raise HTTPException(status_code=404, detail="Gallery link is invalid or expired")
    requested_ids = None
    if payload and payload.media_ids is not None:
        requested_ids = list(dict.fromkeys(payload.media_ids))
        if not requested_ids:
            raise HTTPException(status_code=422, detail="Select at least one photo")
        permitted_ids = set(
            db.scalars(
                select(FaceDetection.photo_id)
                .join(FaceMatch, FaceMatch.detection_id == FaceDetection.id)
                .where(
                    FaceMatch.participant_id == delivery.participant_id,
                    FaceMatch.event_id == delivery.event_id,
                    FaceMatch.state == "high",
                    FaceDetection.photo_id.in_(requested_ids),
                )
                .distinct()
            ).all()
        )
        if permitted_ids != set(requested_ids):
            raise HTTPException(status_code=404, detail="One or more selected photos were not found")
    selection_hash = (
        hashlib.sha256("\n".join(sorted(requested_ids)).encode()).hexdigest() if requested_ids is not None else "ALL"
    )
    existing = db.scalar(
        select(GalleryExport)
        .where(
            GalleryExport.participant_id == delivery.participant_id,
            GalleryExport.event_id == delivery.event_id,
            GalleryExport.selection_hash == selection_hash,
            GalleryExport.status.in_(["QUEUED", "PROCESSING", "READY"]),
            GalleryExport.expires_at > utcnow(),
        )
        .order_by(GalleryExport.created_at.desc())
    )
    if existing:
        return ok({"export_id": existing.id, "status": existing.status}, request)
    job = ProcessingJob(
        organization_id=delivery.organization_id,
        event_id=delivery.event_id,
        job_type="GALLERY_EXPORT",
        status="queued",
        correlation_id=request_id(request),
        max_attempts=5,
    )
    db.add(job)
    db.flush()
    export = GalleryExport(
        organization_id=delivery.organization_id,
        event_id=delivery.event_id,
        participant_id=delivery.participant_id,
        processing_job_id=job.id,
        photo_ids=requested_ids,
        selection_hash=selection_hash,
        expires_at=min(delivery.expires_at, utcnow() + timedelta(hours=24)),
    )
    db.add(export)
    db.flush()
    add_outbox(
        db,
        "fdx.v2.gallery.export.requested",
        "gallery_export",
        export.id,
        {"job_id": job.id, "export_id": export.id},
        delivery.organization_id,
        request_id(request),
    )
    db.commit()
    return ok({"export_id": export.id, "status": export.status}, request)


@router.get("/public/gallery/{token}/exports/{export_id}")
def gallery_export_status(token: str, export_id: str, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "gallery-export-status", token, 60)
    delivery = db.scalar(select(Delivery).where(Delivery.gallery_token_hash == hash_token(token)))
    if (
        not delivery
        or delivery.event.status.upper() in {"DELETION_PENDING", "DELETED", "EXPIRED"}
        or delivery.expires_at <= utcnow()
    ):
        raise HTTPException(status_code=404, detail="Gallery link is invalid or expired")
    item = db.scalar(
        select(GalleryExport).where(
            GalleryExport.id == export_id,
            GalleryExport.event_id == delivery.event_id,
            GalleryExport.participant_id == delivery.participant_id,
        )
    )
    if not item or item.expires_at <= utcnow():
        raise HTTPException(status_code=404, detail="Gallery export was not found or has expired")
    url = (
        storage.presign_get(item.storage_key, filename=f"{delivery.event.name}-photos.zip")
        if item.status == "READY" and item.storage_key
        else None
    )
    if item.status == "READY" and not url:
        url = f"/api/v2/public/gallery/{token}/exports/{item.id}/download"
    return ok(
        {
            "export_id": item.id,
            "status": item.status,
            "size_bytes": item.size_bytes,
            "download_url": url,
            "expires_at": item.expires_at.isoformat(),
            "error": item.error,
        },
        request,
    )


@router.get("/public/gallery/{token}/exports/{export_id}/download")
def download_gallery_export(token: str, export_id: str, request: Request, db: Session = Depends(get_db)):
    check_public_rate_limit(request, "gallery-export-download", token, 20, 300)
    delivery = db.scalar(select(Delivery).where(Delivery.gallery_token_hash == hash_token(token)))
    item = db.scalar(select(GalleryExport).where(GalleryExport.id == export_id))
    if (
        not delivery
        or delivery.event.status.upper() in {"DELETION_PENDING", "DELETED", "EXPIRED"}
        or delivery.expires_at <= utcnow()
        or not item
        or item.participant_id != delivery.participant_id
        or item.event_id != delivery.event_id
        or item.status != "READY"
        or item.expires_at <= utcnow()
        or not item.storage_key
    ):
        raise HTTPException(status_code=404, detail="Gallery export was not found or has expired")
    content, _ = storage.read(item.storage_key)
    filename = (
        "".join(
            character if character.isalnum() or character in "-_" else "-" for character in delivery.event.name
        ).strip("-")
        or "fdx-gallery"
    )
    return Response(
        content=content,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.zip"',
            "Cache-Control": "private, no-store",
        },
    )


def verify_webhook_signature(body: bytes, signature: str | None) -> None:
    if not settings.email_webhook_secret:
        raise HTTPException(status_code=503, detail="Email webhook verification is not configured")
    expected = hmac.new(settings.email_webhook_secret.encode(), body, hashlib.sha256).hexdigest()
    supplied = (signature or "").removeprefix("sha256=")
    if not supplied or not hmac.compare_digest(expected, supplied):
        raise HTTPException(status_code=401, detail="Webhook signature is invalid")


async def process_email_webhook(provider: str, request: Request, signature: str | None, db: Session):
    if settings.email_provider != provider:
        raise HTTPException(status_code=404, detail="Webhook provider is not enabled")
    body = await request.body()
    verify_webhook_signature(body, signature)
    try:
        payload = json.loads(body)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Webhook body is not valid JSON") from exc
    provider_event_id = str(payload.get("id") or payload.get("event_id") or "")
    if not provider_event_id:
        raise HTTPException(status_code=422, detail="Webhook event ID is required")
    if db.scalar(
        select(WebhookEvent.id).where(
            WebhookEvent.provider == provider,
            WebhookEvent.provider_event_id == provider_event_id,
        )
    ):
        return ok({"status": "DUPLICATE"}, request)
    db.add(WebhookEvent(provider=provider, provider_event_id=provider_event_id, payload=payload))
    provider_message_id = (
        payload.get("data", {}).get("email_id") or payload.get("mail", {}).get("messageId") or payload.get("message_id")
    )
    item = (
        db.scalar(select(EmailOutbox).where(EmailOutbox.provider_id == provider_message_id))
        if provider_message_id
        else None
    )
    if item:
        event_type = str(payload.get("type") or payload.get("eventType") or "").lower()
        if any(value in event_type for value in ("delivered", "delivery")):
            item.status = "delivered"
        elif "bounce" in event_type:
            item.status = "bounced"
        elif any(value in event_type for value in ("failed", "complaint", "suppressed")):
            item.status = "failed"
        if item.delivery_id and item.status in {"delivered", "bounced", "failed"}:
            delivery = db.get(Delivery, item.delivery_id)
            if delivery:
                delivery.status = "delivered" if item.status == "delivered" else "failed"
                delivery.participant.delivery_status = delivery.status
    db.commit()
    return ok({"status": "PROCESSED"}, request)


@router.post("/webhooks/email/resend")
async def resend_webhook(
    request: Request,
    x_fdx_webhook_signature: str | None = Header(default=None, alias="X-FDX-Webhook-Signature"),
    db: Session = Depends(get_db),
):
    return await process_email_webhook("resend", request, x_fdx_webhook_signature, db)


@router.post("/webhooks/email/ses")
async def ses_webhook(
    request: Request,
    x_fdx_webhook_signature: str | None = Header(default=None, alias="X-FDX-Webhook-Signature"),
    db: Session = Depends(get_db),
):
    return await process_email_webhook("ses", request, x_fdx_webhook_signature, db)
